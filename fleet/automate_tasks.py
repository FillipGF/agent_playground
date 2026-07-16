#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Скрипт автоматизации обработки задач сервисных инженеров.
Считывает CSV выгрузки, сопоставляет со справочниками из Google Sheets,
сортирует по категориям, генерирует Excel-отчеты с цветовой маркировкой
и рассылает их по соответствующим Telegram-чатам.
Формат Excel полностью соответствует эталонной таблице.
"""

import os
import re
import sys
import time
import random
import json
import math
import glob
import logging
import argparse
import urllib.parse
from datetime import datetime
from io import StringIO

import pandas as pd
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.colors import Color
import gspread
from google.oauth2.service_account import Credentials

# Настройка логирования
# Принудительно устанавливаем кодировку UTF-8 для stdout, чтобы кириллица корректно записывалась в логи
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Справочник порогов заполненности (%) в зависимости от приоритета
PRIORITY_THRESHOLDS = {
    '0': 30,
    'b2b_0': 30,
    'Bronze': 50,
    'b2b_Bronze': 50,
    'Silver': 50,
    'b2b_Silver': 50,
    'Gold': 80,
    'b2b_Gold': 80,
    'Platinum': 80,
    'b2b_Platinum': 80,
    'new': 65,
    'b2b_new': 65
}

# Стили для Excel
gray_color = Color(type='theme', theme=0, tint=-0.15)

FILLS = {
    'Не в сети': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),         # Ярко-красный
    'Пустые': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),            # Ярко-желтый
    'Пополнить': PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid'),         # Ярко-зеленый
    'Разгрузить': PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid'),        # Светло-серый (тема 0, тинт -0.15)
    'Аппараты с ошибками в ячейках': PatternFill(start_color='FFE1BEE7', end_color='FFE1BEE7', fill_type='solid') # Светло-сиреневый
}

# Шрифты для Excel (как в эталоне: Calibri 12)
FONT_CALIBRI_12 = Font(name='Calibri', size=12, bold=False)

def load_config(config_path=None):
    """
    Загружает файл конфигурации.
    """
    if config_path is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.abspath(os.path.join(base_dir, "..", "config.json"))
        
    if not os.path.exists(config_path):
        logger.error(f"Файл конфигурации не найден по пути: {config_path}")
        logger.info("Пожалуйста, создайте config.json на основе шаблона.")
        sys.exit(1)
    
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def download_yandex_fleet_csv(config, park_name, park_id, headless=True):
    """
    Скачивает CSV-файл для конкретного парка Яндекс.Флот с использованием Playwright.
    Сохраняет в inputs/vendings_{park_name}.csv.
    """
    from playwright.sync_api import sync_playwright
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    profile_dir = os.path.abspath(os.path.join(base_dir, ".chrome_profile"))
    inputs_dir = os.path.join(base_dir, "inputs")
    if not os.path.exists(inputs_dir):
        os.makedirs(inputs_dir)
        
    dest_path = os.path.join(inputs_dir, f"vendings_{park_name}.csv")
    url = f"https://fleet.yandex.ru/snickers/vendings?park_id={park_id}"
    
    logger.info(f"Начало скачивания для парка: {park_name} (ID: {park_id})")
    
    max_attempts = 3
    base_delay = 5.0
    
    for attempt in range(1, max_attempts + 1):
        logger.info(f"Попытка {attempt} из {max_attempts} для парка {park_name}...")
        with sync_playwright() as p:
            # Пытаемся запустить браузер с использованием сохраненного профиля
            # Отключаем прокси, чтобы избежать ошибок proxy connection
            context = p.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=headless,
                viewport={"width": 1280, "height": 800},
                args=["--no-proxy-server"]
            )
            
            try:
                page = context.pages[0] if context.pages else context.new_page()
                
                # Переходим на страницу
                logger.info(f"Переход на страницу {url}...")
                page.goto(url)
                
                # Ждем загрузки или редиректа на страницу авторизации
                page.wait_for_timeout(5000)
                
                if "passport.yandex" in page.url:
                    logger.error(f"Ошибка: Необходима повторная авторизация в Яндекс для парка {park_name}!")
                    logger.error("Пожалуйста, запустите скрипт 'login_and_inspect.py' вручную для входа.")
                    raise Exception("Необходима авторизация")
                    
                # Ждем появления кнопки экспорта
                logger.info("Ожидание кнопки экспорта в CSV...")
                export_button_selector = 'button[aria-label="Кнопка экспорта в CSV"]'
                
                try:
                    page.wait_for_selector(export_button_selector, timeout=20000)
                except Exception:
                    # Если кнопка не найдена, проверим не улетели ли мы на авторизацию
                    if "passport.yandex" in page.url:
                        logger.error(f"Ошибка: Необходима авторизация в Яндекс для парка {park_name}!")
                        raise Exception("Необходима авторизация")
                    else:
                        logger.error(f"Не удалось дождаться кнопки экспорта для парка {park_name}!")
                        debug_tools_dir = os.path.abspath(os.path.join(base_dir, "..", "debug_tools"))
                        os.makedirs(debug_tools_dir, exist_ok=True)
                        screenshot_path = os.path.join(debug_tools_dir, f"error_{park_name}_attempt_{attempt}.png")
                        page.screenshot(path=screenshot_path)
                        logger.info(f"Скриншот ошибки сохранен в {screenshot_path}")
                        raise Exception("Кнопка экспорта не найдена")
                
                # Кликаем на кнопку экспорта и ждем начала скачивания
                logger.info("Клик по кнопке экспорта и скачивание файла...")
                with page.expect_download(timeout=30000) as download_info:
                    page.locator(export_button_selector).click()
                    
                download = download_info.value
                download.save_as(dest_path)
                logger.info(f"Файл успешно скачан и сохранен в: {dest_path}")
                return dest_path
                
            except Exception as e:
                logger.error(f"Исключение на попытке {attempt} при скачивании парка {park_name}: {e}")
                if attempt < max_attempts:
                    delay = base_delay * attempt + random.uniform(1.0, 3.0)
                    logger.info(f"Ожидание {delay:.2f} сек перед повторной попыткой...")
                    time.sleep(delay)
                else:
                    raise e
            finally:
                context.close()

def download_yandex_map_service_csv(config, park_name, park_id, headless=True):
    """
    Скачивает CSV-файл сервисной карты для конкретного парка Яндекс.Флот с использованием Playwright.
    Сохраняет в inputs/map_service_{park_name}.csv.
    """
    from playwright.sync_api import sync_playwright
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    profile_dir = os.path.abspath(os.path.join(base_dir, ".chrome_profile"))
    inputs_dir = os.path.join(base_dir, "inputs")
    if not os.path.exists(inputs_dir):
        os.makedirs(inputs_dir)
        
    dest_path = os.path.join(inputs_dir, f"map_service_{park_name}.csv")
    url = f"https://fleet.yandex.ru/snickers/map/service?park_id={park_id}&location_warehouse=false"
    
    logger.info(f"Начало скачивания сервисной карты для парка: {park_name} (ID: {park_id})")
    
    max_attempts = 3
    base_delay = 5.0
    
    for attempt in range(1, max_attempts + 1):
        logger.info(f"Попытка {attempt} из {max_attempts} для парка {park_name}...")
        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=headless,
                viewport={"width": 1400, "height": 900},
                args=["--no-proxy-server"]
            )
            
            try:
                page = context.pages[0] if context.pages else context.new_page()
                
                logger.info(f"Переход на страницу {url}...")
                page.goto(url)
                page.wait_for_timeout(10000) # Ждем загрузки API
                
                if "passport.yandex" in page.url:
                    logger.error(f"Ошибка: Необходима повторная авторизация в Яндекс для парка {park_name}!")
                    raise Exception("Необходима авторизация")
                    
                # Ищем кнопку экспорта
                logger.info("Ожидание кнопки экспорта в CSV...")
                export_button_selector = 'button[aria-label*="CSV"]'
                
                try:
                    page.wait_for_selector(export_button_selector, timeout=20000)
                except Exception:
                    if "passport.yandex" in page.url:
                        logger.error(f"Ошибка: Необходима авторизация в Яндекс для парка {park_name}!")
                        raise Exception("Необходима авторизация")
                    else:
                        logger.error(f"Не удалось дождаться кнопки экспорта для парка {park_name}!")
                        debug_tools_dir = os.path.abspath(os.path.join(base_dir, "..", "debug_tools"))
                        os.makedirs(debug_tools_dir, exist_ok=True)
                        screenshot_path = os.path.join(debug_tools_dir, f"error_map_{park_name}_attempt_{attempt}.png")
                        page.screenshot(path=screenshot_path)
                        logger.info(f"Скриншот ошибки сохранен в {screenshot_path}")
                        raise Exception("Кнопка экспорта не найдена")
                
                logger.info("Клик по кнопке экспорта и скачивание файла...")
                with page.expect_download(timeout=30000) as download_info:
                    page.locator(export_button_selector).first.click()
                    
                download = download_info.value
                download.save_as(dest_path)
                logger.info(f"Файл успешно скачан и сохранен в: {dest_path}")
                return dest_path
                
            except Exception as e:
                logger.error(f"Исключение на попытке {attempt} при скачивании карты {park_name}: {e}")
                if attempt < max_attempts:
                    delay = base_delay * attempt + random.uniform(1.0, 3.0)
                    logger.info(f"Ожидание {delay:.2f} сек перед повторной попыткой...")
                    time.sleep(delay)
                else:
                    raise e
            finally:
                context.close()

def load_map_service_norms(city, inputs_dir):
    """
    Загружает нормы заполненности из единого файла кэша fullness_norms.json.
    Возвращает словарь {vending_id: threshold_percent}.
    """
    file_path = os.path.join(inputs_dir, "fullness_norms.json")
    if not os.path.exists(file_path):
        logger.warning(f"Файл кэша норм заполненности не найден: {file_path}")
        return {}
        
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            norms = json.load(f)
        logger.info(f"Успешно загружен кэш норм заполненности из JSON. Записей: {len(norms)}")
        return norms
    except Exception as e:
        logger.error(f"Ошибка при чтении файла кэша норм заполненности {file_path}: {e}")
        return {}


def load_empty_stations(inputs_dir):
    """
    Загружает кэш истории пустых станций из empty_stations.json.
    Возвращает словарь {vending_id: first_seen_iso_timestamp}.
    """
    file_path = os.path.join(inputs_dir, "empty_stations.json")
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Не удалось загрузить историю пустых станций: {e}")
        return {}


def save_empty_stations(inputs_dir, data):
    """
    Сохраняет кэш истории пустых станций в empty_stations.json.
    """
    file_path = os.path.join(inputs_dir, "empty_stations.json")
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Не удалось сохранить историю пустых станций: {e}")


def fetch_google_sheet(sheet_id, sheet_name):
    """
    Загружает указанный лист из Google Sheets без OAuth-авторизации
    через Google Visualization API в формате CSV.
    
    :param sheet_id: str - ID таблицы
    :param sheet_name: str - имя листа
    :return: DataFrame
    """
    encoded_name = urllib.parse.quote(sheet_name)
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={encoded_name}"
    logger.info(f"Запрос справочника '{sheet_name}'...")
    try:
        response = requests.get(url, timeout=15)
        if response.status_code == 200:
            df = pd.read_csv(StringIO(response.text))
            logger.info(f"Справочник '{sheet_name}' успешно загружен. Строк: {len(df)}")
            return df
        else:
            logger.error(f"Не удалось загрузить '{sheet_name}': HTTP status {response.status_code}")
    except Exception as e:
        logger.error(f"Ошибка при загрузке '{sheet_name}': {e}")
    return None

def format_phone_number(raw_phone):
    """
    Приводит телефонные номера к единой маске РФ (7XXXXXXXXXX).
    
    :param raw_phone: str или float - сырой номер телефона
    :return: str
    """
    if pd.isna(raw_phone):
        return ""
    
    raw_phone_str = str(raw_phone).strip()
    if not raw_phone_str:
        return ""

    # Ищем паттерны номеров телефонов с помощью регулярного выражения
    # Аналог JS: /(?<!\d)(?:(?:\+7|8|7)\D*)?(?:\d\D*){10}(?!\d)/g
    pattern = re.compile(r'(?<!\d)(?:(?:\+7|8|7)\D*)?(?:\d\D*){10}(?!\d)')
    matches = pattern.findall(raw_phone_str)

    if not matches:
        return raw_phone_str

    formatted_numbers = []
    for match in matches:
        # Очищаем от всех нецифровых символов
        clean = re.sub(r'\D', '', match)
        
        # Приводим к стандарту РФ
        if len(clean) == 11 and clean.startswith('8'):
            clean = '7' + clean[1:]
        elif len(clean) == 10:
            clean = '7' + clean
            
        formatted_numbers.append(clean)

    return ", ".join(formatted_numbers)

def fix_address(raw_address):
    """
    Исправляет ошибки дат, вызванные авто-форматированием Excel.
    
    :param raw_address: str - адрес для исправления
    :return: str
    """
    if pd.isna(raw_address):
        return ""
    
    addr_str = str(raw_address).strip()
    
    # 1. Заменяем июнь 65-го года в любых его проявлениях на номер дома "65/6"
    # JS: str.replace(/0?6\/0?1\/(?:19)?65(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?/g, '65/6')
    addr_str = re.sub(r'0?6/0?1/(?:19)?65(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?', '65/6', addr_str)

    # 2. Универсальный фикс для дробных домов вида Месяц/1/Год -> Год/Месяц (например, 12/5)
    # JS: str.replace(/0?(\d{1,2})\/0?1\/(?:19|20)?(\d{2})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?/g, '$2/$1')
    addr_str = re.sub(r'0?(\d{1,2})/0?1/(?:19|20)?(\d{2})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?', r'\2/\1', addr_str)

    return addr_str

def is_within_working_hours(local_time, open_time_str, close_time_str):
    """
    Проверяет, входит ли текущее локальное время в рабочий интервал точки.
    
    :param local_time: datetime - текущее локальное время города
    :param open_time_str: str - время открытия (например, "08:00")
    :param close_time_str: str - время закрытия (например, "22:00")
    :return: bool
    """
    import pandas as pd
    if pd.isna(open_time_str) or pd.isna(close_time_str):
        return True
        
    open_str = str(open_time_str).strip()[:5]
    close_str = str(close_time_str).strip()[:5]
    
    if not open_str or not close_str or open_str == 'nan' or close_str == 'nan':
        return True
        
    try:
        from datetime import time
        
        # Парсим часы и минуты
        open_h, open_m = map(int, open_str.split(':'))
        close_h, close_m = map(int, close_str.split(':'))
        
        t_open = time(open_h, open_m)
        t_close = time(close_h, close_m)
        t_curr = local_time.time()
        
        if t_open <= t_close:
            # Обычный дневной режим (например, 08:00 - 22:00)
            return t_open <= t_curr <= t_close
        else:
            # Ночной режим работы с переходом через полночь (например, 15:00 - 03:00)
            return t_curr >= t_open or t_curr <= t_close
    except Exception:
        # Если формат времени некорректный, считаем точку работающей
        return True

def apply_filtering_logic(df, priority_guide, appointment_person, main_city="Неизвестно", map_service_norms=None, empty_stations_history=None):
    """
    Применяет бизнес-логику фильтрации и категоризации задач к датафрейму.
    Возвращает список словарей, готовых к генерации Excel.
    
    :param df: DataFrame - исходные данные
    :param priority_guide: dict - справочник приоритетов
    :param appointment_person: dict - справочник ответственных инженеров
    :param main_city: str - основной город этого файла для определения таймзоны
    :param map_service_norms: dict - словарь норм заполненности из сервисной карты (vending_id -> threshold)
    :param empty_stations_history: dict - история пустых станций {vending_id: first_seen_iso}
    :return: list of dict
    """
    records = []

    # Разница временных поясов от МСК (ключи в нижнем регистре с заменой 'ё' -> 'е')
    CITY_OFFSETS = {
        "омск": 3,
        "магнитогорск": 2,
        "сургут": 2,
        "ижевск": 1,
        "ульяновск": 1,
        "рязань": 0,
        "киров": 0,
        "чебоксары": 0,
        "орел": 0
    }

    from datetime import datetime, timezone, timedelta
    
    # Определяем текущее локальное время в городе
    now_msk = datetime.now(timezone(timedelta(hours=3)))
    city_key = str(main_city).strip().lower().replace("ё", "е")
    offset = CITY_OFFSETS.get(city_key, 0)
    local_time = now_msk + timedelta(hours=offset)
    local_time_naive = local_time.replace(tzinfo=None)
    
    logger.info(
        f"Часовой пояс для '{main_city}': +{offset} ч. от МСК. "
        f"Время МСК: {now_msk.strftime('%H:%M')}, "
        f"Локальное время: {local_time.strftime('%H:%M')}"
    )

    # 1. Фильтрация пустых или невалидных строк
    df_cleaned = df.dropna(subset=['DisplayNumber']).copy()
    
    # Преобразуем DisplayNumber в строку без плавающей точки, если она есть
    df_cleaned['DisplayNumber_str'] = df_cleaned['DisplayNumber'].apply(
        lambda x: str(int(float(x))).strip() if pd.notna(x) else ""
    )
    
    for _, row in df_cleaned.iterrows():
        place_name = str(row.get('PlaceName', ''))
        
        # Фильтр: Исключаем Офисы, Склады, Регионы, ИП Барсуков, почту/собаки
        if (
            "Офис" in place_name or
            "@" in place_name or
            "Регионы." in place_name or
            "ИП Барсуков" in place_name
        ):
            continue

        display_number = row['DisplayNumber_str']
        
        # Получаем приоритет (по умолчанию 'new')
        priority = priority_guide.get(display_number, 'new')
        if pd.isna(priority) or not str(priority).strip():
            priority = 'new'
        else:
            priority = str(priority).strip()

        # Получаем ответственного и форматируем его телефон
        responsible = appointment_person.get(display_number, '')
        responsible_formatted = format_phone_number(responsible)

        # Выбираем адрес (ищем наиболее полную колонку, содержащую 'Address')
        address_cols = [col for col in df.columns if 'Address' in col]
        actual_address = ""
        for col in address_cols:
            val = str(row.get(col, '')).strip()
            if len(val) > len(actual_address):
                actual_address = val
        
        clean_address = fix_address(actual_address)

        # Собираем базовую информацию
        base_record = {
            'DisplayNumber': display_number,
            'PlaceName': str(row.get('PlaceName', '')) if pd.notna(row.get('PlaceName')) else '',
            'Address': clean_address,
            'LocationOpenTime': str(row.get('LocationOpenTime', '')) if pd.notna(row.get('LocationOpenTime')) else '',
            'LocationCloseTime': str(row.get('LocationCloseTime', '')) if pd.notna(row.get('LocationCloseTime')) else '',
            'Приоритет': priority,
            'Отвественный': responsible_formatted,
            'City': str(row.get('City', 'Неизвестно')) if pd.notna(row.get('City')) else 'Неизвестно',
            'Comment': ''
        }

        status = str(row.get('Status', '')).lower().strip()

        # Категория 1: Не в сети
        if status == 'not_responding':
            # Проверяем рабочее время точки
            open_time = row.get('LocationOpenTime')
            close_time = row.get('LocationCloseTime')
            
            # Если сейчас не рабочее время, пропускаем создание задачи "Не в сети"
            if not is_within_working_hours(local_time, open_time, close_time):
                continue
                
            rec = base_record.copy()
            rec['Category'] = 'Не в сети'
            records.append(rec)
            continue

        # Категория 2: Работающие автоматы (ok или пусто)
        if status in ['ok', '']:
            try:
                total = float(row.get('CellsTotal', 0))
                empty = float(row.get('FreeCellsCount', 0))
            except (ValueError, TypeError):
                continue

            if math.isnan(total) or total == 0:
                continue

            filled = total - empty
            fill_percent = (filled / total) * 100
            
            # Определяем порог: сначала ищем на сервисной карте, затем в PRIORITY_THRESHOLDS
            if map_service_norms and display_number in map_service_norms:
                threshold = map_service_norms[display_number]
            else:
                threshold = PRIORITY_THRESHOLDS.get(priority, 65)
                
            exact_target = (total * threshold) / 100
            
            target_unload = math.ceil(exact_target)
            target_replenish = round(exact_target)

            # ПУСТЫЕ (0 банок)
            if total == empty:
                rec = base_record.copy()
                rec['Category'] = 'Пустые'
                if empty_stations_history and display_number in empty_stations_history:
                    try:
                        from datetime import datetime
                        first_seen_str = empty_stations_history[display_number]
                        first_seen = datetime.fromisoformat(first_seen_str).replace(tzinfo=None)
                        delta_hours = (local_time_naive - first_seen).total_seconds() / 3600.0
                        if delta_hours >= 48.0:
                            first_seen_formatted = first_seen.strftime("%d.%m %H:%M")
                            rec['Comment'] = f"[!] Пустой более 48ч (с {first_seen_formatted})"
                    except Exception as history_err:
                        logger.error(f"Ошибка проверки истории пустой станции {display_number}: {history_err}")
                records.append(rec)
            
            # РАЗГРУЗИТЬ (filled - targetUnload >= 2)
            elif (filled - target_unload) >= 2:
                rec = base_record.copy()
                rec['Category'] = 'Разгрузить'
                records.append(rec)

            # ПОПОЛНИТЬ
            elif (
                fill_percent <= threshold and
                (target_replenish - filled) >= 2 and
                empty >= 2
            ):
                rec = base_record.copy()
                rec['Category'] = 'Пополнить'
                records.append(rec)

            # Ошибки в ячейках (проверяются параллельно)
            error_cells = str(row.get('Error cells', '[]')).strip()
            if error_cells and error_cells != '[]' and error_cells != 'nan':
                rec = base_record.copy()
                rec['Category'] = 'Аппараты с ошибками в ячейках'
                records.append(rec)

    return records

def save_styled_excel(records, output_path):
    """
    Генерирует красиво отформатированный Excel-файл на основе записей.
    Формат полностью соответствует эталону: без заголовков колонок,
    с разделением строк по категориям-заголовкам и заливкой только колонок 2-6 (B-F).
    
    :param records: list of dict - записи задач
    :param output_path: str - путь к сохранению файла
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результат"
    
    # Включаем отображение сетки
    ws.views.sheetView[0].showGridLines = True

    # Порядок выгрузки категорий
    categories_order = [
        'Не в сети',
        'Пустые',
        'Пополнить',
        'Разгрузить',
        'Аппараты с ошибками в ячейках'
    ]

    current_row = 1
    for category in categories_order:
        # Фильтруем записи по текущей категории
        cat_records = [r for r in records if r.get('Category') == category]
        if not cat_records:
            continue
            
        # Записываем строку-заголовок категории (Col A имеет название категории, остальные - None)
        ws.append([category, None, None, None, None, None, None, None])
        header_cell = ws.cell(row=current_row, column=1)
        header_cell.font = FONT_CALIBRI_12
        header_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Сбрасываем форматирование (границы и заливку) для всей строки заголовка категории
        for c in range(1, 9):
            cell = ws.cell(row=current_row, column=c)
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            
        current_row += 1
        
        # Записываем строки данных
        for r in cat_records:
            row_vals = [
                r.get('DisplayNumber', ''),
                r.get('PlaceName', ''),
                r.get('Address', ''),
                r.get('LocationOpenTime', ''),
                r.get('LocationCloseTime', ''),
                r.get('Приоритет', ''),
                r.get('Отвественный', ''),
                r.get('Comment', '')
            ]
            ws.append(row_vals)
            
            row_fill = FILLS.get(category, None)
            
            for col_idx in range(1, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = FONT_CALIBRI_12
                cell.border = Border() # Без рамок (как в эталоне)
                cell.alignment = Alignment(horizontal='general', vertical='center')
                
                # Заливку применяем только к колонкам 2-6 (B, C, D, E, F)
                if col_idx in [2, 3, 4, 5, 6] and row_fill:
                    cell.fill = row_fill
                else:
                    cell.fill = PatternFill(fill_type=None)
                    
            current_row += 1

    # Задаем фиксированную ширину колонок в соответствии с эталоном
    ws.column_dimensions['A'].width = 13.0
    ws.column_dimensions['B'].width = 69.125
    ws.column_dimensions['C'].width = 111.125
    ws.column_dimensions['D'].width = 13.0
    ws.column_dimensions['E'].width = 13.0
    ws.column_dimensions['F'].width = 13.0
    ws.column_dimensions['G'].width = 13.0
    ws.column_dimensions['H'].width = 30.0

    wb.save(output_path)
    logger.info(f"Файл успешно сохранен по эталонному формату: {output_path}")

# Цвета для Google Sheets (RGB от 0.0 до 1.0)
GS_COLORS = {
    'Не в сети': {'red': 1.0, 'green': 0.0, 'blue': 0.0},                  # Ярко-красный
    'Пустые': {'red': 1.0, 'green': 1.0, 'blue': 0.0},                     # Ярко-желтый
    'Пополнить': {'red': 146/255.0, 'green': 208/255.0, 'blue': 80/255.0},  # Зеленый
    'Разгрузить': {'red': 217/255.0, 'green': 217/255.0, 'blue': 217/255.0}, # Светло-серый
    'Аппараты с ошибками в ячейках': {'red': 225/255.0, 'green': 190/255.0, 'blue': 231/255.0} # Сиреневый
}

def safe_gspread_call(func, *args, **kwargs):
    """
    Выполняет вызов Google Sheets API с обработкой ошибок превышения квоты (429)
    и повторными попытками с экспоненциальной задержкой.
    """
    max_retries = 5
    base_delay = 2.0
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            is_quota_error = False
            try:
                if hasattr(e, 'response') and e.response is not None:
                    if e.response.status_code == 429:
                        is_quota_error = True
            except Exception:
                pass
            
            if not is_quota_error and ("429" in str(e) or "quota" in str(e).lower()):
                is_quota_error = True
                
            if is_quota_error and attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt) + random.uniform(0.5, 1.5)
                func_name = getattr(func, '__name__', str(func))
                logger.warning(
                    f"Превышена квота Google Sheets API (429). Повтор запроса '{func_name}' "
                    f"через {delay:.2f} сек (попытка {attempt + 1}/{max_retries})..."
                )
                time.sleep(delay)
            else:
                raise e

def authenticate_gspread(config):
    """
    Выполняет аутентификацию в Google Sheets с использованием сервисного аккаунта.
    Если файл ключа не найден, выводит подробную инструкцию и возвращает None.
    """
    credentials_file = config.get("google_service_account_file", "service_account.json")
    if not os.path.isabs(credentials_file):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        credentials_file = os.path.abspath(os.path.join(base_dir, "..", credentials_file))
    if not os.path.exists(credentials_file):
        logger.warning("=" * 80)
        logger.warning("ПРЕДУПРЕЖДЕНИЕ: Файл авторизации Google Service Account не найден!")
        logger.warning(f"Ожидаемый путь: {os.path.abspath(credentials_file)}")
        logger.warning("Для автоматического обновления Google Таблиц:")
        logger.warning("  1. Создайте проект в Google Cloud Console (https://console.cloud.google.com/).")
        logger.warning("  2. Включите Google Sheets API и Google Drive API.")
        logger.warning("  3. Создайте Service Account и скачайте JSON-ключ (Key -> Add Key -> Create new key -> JSON).")
        logger.warning(f"  4. Переименуйте ключ в '{credentials_file}' и положите его в корень рабочей папки.")
        logger.warning("  5. Поделитесь вашими Google Таблицами (Editor) с email-адресом сервисного аккаунта (он указан в JSON-файле в поле client_email).")
        logger.warning("=" * 80)
        return None

    try:
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(credentials_file, scopes=scopes)
        client = gspread.authorize(creds)
        logger.info("Успешная авторизация в Google Sheets API.")
        return client
    except Exception as e:
        logger.error(f"Ошибка при авторизации в Google Sheets API: {e}")
        return None

def col_to_a1(col_idx):
    """
    Преобразует 1-индексный номер колонки в буквенное обозначение Excel/Google Sheets (например, 1 -> A, 27 -> AA).
    """
    letters = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

def set_google_sheet_column_widths(spreadsheet, worksheet_id, widths_dict):
    """
    Устанавливает ширину колонок в пикселях через batch_update.
    widths_dict: словарь, где ключи - 0-индексные номера колонок (int), а значения - ширина в пикселях (int).
    """
    requests_list = []
    for col_idx, width in widths_dict.items():
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": worksheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": col_idx,
                    "endIndex": col_idx + 1
                },
                "properties": {
                    "pixelSize": width
                },
                "fields": "pixelSize"
            }
        })
    if requests_list:
        safe_gspread_call(spreadsheet.batch_update, {"requests": requests_list})

def update_google_sheet_city(client, sheet_id, city_name, records):
    """
    Полностью перезаписывает лист с именем города в Google Таблице с цветовой разметкой.
    """
    logger.info(f"Обновление Google Таблицы задач ({city_name})...")
    try:
        # Открываем таблицу с обработкой ошибок
        spreadsheet = safe_gspread_call(client.open_by_key, sheet_id)
        
        # Ищем лист города, если нет — создаем
        try:
            worksheet = safe_gspread_call(spreadsheet.worksheet, city_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = safe_gspread_call(spreadsheet.add_worksheet, title=city_name, rows=1000, cols=20)
            logger.info(f"Создан новый лист: '{city_name}'")
            
        # Сначала очистим лист (значения и форматирование)
        safe_gspread_call(worksheet.clear)
        
        # Порядок категорий
        categories_order = [
            'Не в сети',
            'Пустые',
            'Пополнить',
            'Разгрузить',
            'Аппараты с ошибками в ячейках'
        ]
        
        rows_to_write = []
        color_ranges = []
        
        current_row = 1
        for category in categories_order:
            cat_records = [r for r in records if r.get('Category') == category]
            if not cat_records:
                continue
                
            # Записываем строку-заголовок категории (Col A имеет название категории, остальные - пустые)
            rows_to_write.append([category, "", "", "", "", "", "", ""])
            current_row += 1
            
            start_row = current_row
            for r in cat_records:
                rows_to_write.append([
                    r.get('DisplayNumber', ''),
                    r.get('PlaceName', ''),
                    r.get('Address', ''),
                    r.get('LocationOpenTime', ''),
                    r.get('LocationCloseTime', ''),
                    r.get('Приоритет', ''),
                    r.get('Отвественный', ''),
                    r.get('Comment', '')
                ])
                current_row += 1
            end_row = current_row - 1
            
            # Сохраняем диапазон для раскраски
            if start_row <= end_row:
                color_ranges.append((start_row, end_row, category))
                
        if not rows_to_write:
            logger.warning("Нет данных для записи в Google Таблицу.")
            return
            
        # Обновляем все значения за один раз
        safe_gspread_call(worksheet.update, values=rows_to_write)
        
        # Формируем батч запросов на форматирование
        last_row = len(rows_to_write)
        formats = [
            {
                "range": f"A1:H{last_row}",
                "format": {
                    "textFormat": {
                        "fontFamily": "Calibri",
                        "fontSize": 12,
                        "bold": False
                    },
                    "verticalAlignment": "MIDDLE"
                }
            }
        ]
        
        # Добавляем раскраску B-F для каждой категории
        for start_r, end_r, cat in color_ranges:
            color = GS_COLORS.get(cat)
            if color:
                formats.append({
                    "range": f"B{start_r}:F{end_r}",
                    "format": {
                        "backgroundColor": color
                    }
                })
                
        # Применяем форматирование одним батчем
        safe_gspread_call(worksheet.batch_format, formats)
        
        # Устанавливаем ширину колонок в Google Sheets
        widths = {
            0: 100,  # A
            1: 300,  # B
            2: 300,  # C
            3: 100,  # D
            4: 100,  # E
            5: 100,  # F
            6: 100,  # G
            7: 200   # H
        }
        set_google_sheet_column_widths(spreadsheet, worksheet.id, widths)
            
        logger.info(f"Google Таблица задач ({city_name}) успешно обновлена.")
    except Exception as e:
        logger.error(f"Не удалось обновить Google Таблицу задач ({city_name}): {e}")

def update_google_sheet_disconnected(client, sheet_id, city_name, records):
    """
    Добавляет отключенные автоматы ("Не в сети") в Google Таблицу отключенных станций
    новым блоком из 8 колонок справа.
    """
    # Фильтруем задачи "Не в сети"
    offline_tasks = [r for r in records if r.get('Category') == 'Не в сети']
    if not offline_tasks:
        logger.info(f"Нет задач 'Не в сети' для записи в Google Таблицу отключенных станций ({city_name}).")
        return

    logger.info(f"Добавление {len(offline_tasks)} отключенных станций в Google Таблицу отключенных станций ({city_name})...")
    try:
        spreadsheet = safe_gspread_call(client.open_by_key, sheet_id)
        
        # Ищем лист города, если нет — создаем
        try:
            worksheet = safe_gspread_call(spreadsheet.worksheet, city_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = safe_gspread_call(spreadsheet.add_worksheet, title=city_name, rows=1000, cols=20)
            logger.info(f"Создан новый лист для отключенных станций: '{city_name}'")
        
        # Сначала получаем первую строку для поиска первого блока с датой
        row_1 = safe_gspread_call(worksheet.row_values, 1)
        
        # Функция парсинга даты из строки
        def parse_date_str(val_str):
            match = re.match(r'^(\d{1,2})[./](\d{1,2})[./](\d{4})$', val_str)
            if match:
                return (int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return None

        # Разница временных поясов от МСК
        CITY_OFFSETS = {
            "омск": 3,
            "магнитогорск": 2,
            "сургут": 2,
            "ижевск": 1,
            "ульяновск": 1,
            "рязань": 0,
            "киров": 0,
            "чебоксары": 0,
            "орел": 0
        }
        from datetime import timezone, timedelta
        now_msk = datetime.now(timezone(timedelta(hours=3)))
        city_key = str(city_name).strip().lower().replace("ё", "е")
        offset = CITY_OFFSETS.get(city_key, 0)
        local_time = now_msk + timedelta(hours=offset)
        
        date_str = local_time.strftime("%d.%m.%Y")
        today_tuple = (local_time.day, local_time.month, local_time.year)
        
        if now_msk.hour < 12:
            time_of_day = "Утро"
        elif now_msk.hour < 17:
            time_of_day = "День"
        else:
            time_of_day = "Вечер"
        
        # Находим все индексы, где указаны даты в первой строке
        date_indices = []
        for idx, val in enumerate(row_1):
            val_str = str(val).strip()
            if val_str and any(char.isdigit() for char in val_str) and ('.' in val_str or '/' in val_str):
                date_indices.append(idx)

        # Проверяем, существует ли сегодняшний блок (он должен быть самым правым, т.е. date_indices[-1])
        today_block_exists = False
        if date_indices:
            last_date_val = str(row_1[date_indices[-1]]).strip()
            if parse_date_str(last_date_val) == today_tuple:
                today_block_exists = True

        new_rows = []
        for r in offline_tasks:
            new_rows.append([
                r.get('DisplayNumber', ''),
                r.get('PlaceName', ''),
                r.get('Address', ''),
                r.get('LocationOpenTime', ''),
                r.get('LocationCloseTime', ''),
                r.get('Приоритет', ''),
                r.get('Отвественный', ''),
                ''
            ])

        sections = {}
        previous_row_count = 0
        
        if today_block_exists:
            target_col_start = date_indices[-1]
            logger.info(f"Обнаружен сегодняшний блок для г. {city_name} в колонке {col_to_a1(target_col_start + 1)}. Будем добавлять/обновлять отчет...")
            
            # Читаем значения первой колонки сегодняшнего блока, чтобы определить свободные строки
            col_vals = safe_gspread_call(worksheet.col_values, target_col_start + 1)
            previous_row_count = len(col_vals)
            
            # Читаем все значения сегодняшнего блока
            start_col_letter = col_to_a1(target_col_start + 1)
            end_col_letter = col_to_a1(target_col_start + 8)
            range_name = f"{start_col_letter}1:{end_col_letter}{previous_row_count}"
            existing_data = safe_gspread_call(worksheet.get, range_name)
            
            # Парсим существующие секции
            current_section_label = None
            current_section_rows = []
            for row in existing_data:
                if not row:
                    continue
                first_val = str(row[0]).strip() if len(row) > 0 else ""
                second_val = str(row[1]).strip() if len(row) > 1 else ""
                
                if parse_date_str(first_val) == today_tuple and second_val in ["Утро", "День", "Вечер"]:
                    if current_section_label:
                        sections[current_section_label] = current_section_rows
                    current_section_label = second_val
                    current_section_rows = []
                else:
                    if current_section_label:
                        padded_row = row + [""] * (8 - len(row))
                        if str(padded_row[0]).strip():
                            current_section_rows.append(padded_row)
            if current_section_label:
                sections[current_section_label] = current_section_rows
            
            # Обновляем/добавляем текущую секцию
            sections[time_of_day] = new_rows
            
        else:
            # Первый запуск за день -> Удаляем самый старый блок для освобождения места и пишем в новый столбец
            logger.info(f"Сегодняшний блок для г. {city_name} не найден. Проводим удаление старейшего отчета...")
            
            # Находим первый индекс, где указана дата (старейший блок)
            first_date_col = None
            for idx, val in enumerate(row_1):
                val_str = str(val).strip()
                if val_str and any(char.isdigit() for char in val_str) and ('.' in val_str or '/' in val_str):
                    first_date_col = idx + 1
                    break
                    
            # Если дата найдена, удаляем блок из 8 колонок
            if first_date_col is not None:
                logger.info(f"Удаление старейшего блока из 8 колонок (начиная с колонки {col_to_a1(first_date_col)}) на листе {city_name}...")
                safe_gspread_call(worksheet.delete_columns, first_date_col, first_date_col + 7)
            else:
                # Fallback: если дат нет, удаляем первые 8 колонок по умолчанию
                if worksheet.col_count >= 9:
                    logger.info(f"Даты не найдены. Удаление первых 8 колонок по умолчанию (B:I) на листе {city_name}...")
                    safe_gspread_call(worksheet.delete_columns, 2, 9)
            
            # Перечитываем первую строку, так как столбцы сдвинулись
            row_1 = safe_gspread_call(worksheet.row_values, 1)
            
            # Ищем новый крайний блок
            date_indices = []
            for idx, val in enumerate(row_1):
                val_str = str(val).strip()
                if val_str and any(char.isdigit() for char in val_str) and ('.' in val_str or '/' in val_str):
                    date_indices.append(idx)
                    
            if date_indices:
                target_col_start = date_indices[-1] + 8
            else:
                target_col_start = 1
                
            sections = {time_of_day: new_rows}
            
            needed_cols = target_col_start + 8
            if needed_cols > worksheet.col_count:
                cols_to_add = needed_cols - worksheet.col_count
                safe_gspread_call(worksheet.add_cols, cols_to_add)
                logger.info(f"Добавлено {cols_to_add} колонок в Google Таблицу.")

        # Строим block_data
        block_data = []
        for label in ["Утро", "День", "Вечер"]:
            if label in sections:
                is_first_section = (len(block_data) == 0)
                comment_val = "Комментарий" if is_first_section else ""
                header_row = [date_str, label, "", "", "", "", comment_val, ""]
                block_data.append(header_row)
                block_data.extend(sections[label])
                block_data.append([""] * 8)
                
        # Если новый блок короче предыдущего, дополняем пустыми строками
        if previous_row_count > 0 and len(block_data) < previous_row_count:
            padding_len = previous_row_count - len(block_data)
            block_data.extend([[""] * 8] * padding_len)
            
        start_row = 1
        needed_rows = start_row + len(block_data) - 1
        if needed_rows > worksheet.row_count:
            rows_to_add = needed_rows - worksheet.row_count
            safe_gspread_call(worksheet.add_rows, rows_to_add)
            logger.info(f"Добавлено {rows_to_add} строк в Google Таблицу.")
            
        # Подготовим диапазон для записи
        start_col_letter = col_to_a1(target_col_start + 1)
        end_col_letter = col_to_a1(target_col_start + 8)
        range_name = f"{start_col_letter}{start_row}:{end_col_letter}{needed_rows}"
        
        # Записываем данные
        safe_gspread_call(worksheet.update, values=block_data, range_name=range_name)
        
        # Строим formats для пакетного обновления
        formats = [
            {
                "range": range_name,
                "format": {
                    "textFormat": {
                        "fontFamily": "Calibri",
                        "fontSize": 12,
                        "bold": False
                     },
                    "verticalAlignment": "MIDDLE",
                    "backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}
                }
            }
        ]
        
        # Заливка для "Не в сети" (красный) для колонок PlaceName, Address, OpenTime, CloseTime, Priority
        color_start_col = col_to_a1(target_col_start + 2)
        color_end_col = col_to_a1(target_col_start + 6)
        
        current_offset = 0
        for label in ["Утро", "День", "Вечер"]:
            if label in sections:
                section_tasks_count = len(sections[label])
                if section_tasks_count > 0:
                    start_r_in_sheet = start_row + current_offset + 1
                    end_r_in_sheet = start_row + current_offset + section_tasks_count
                    
                    formats.append({
                        "range": f"{color_start_col}{start_r_in_sheet}:{color_end_col}{end_r_in_sheet}",
                        "format": {
                            "backgroundColor": GS_COLORS['Не в сети']
                        }
                    })
                current_offset += 1 + section_tasks_count + 1
                
        safe_gspread_call(worksheet.batch_format, formats)
        
        # Устанавливаем ширину колонок для нового блока (только если создавали новый столбец)
        if not today_block_exists:
            widths_dict = {}
            for offset_col, w in enumerate([100, 500, 800, 100, 100, 100, 150, 50]):
                col_idx = target_col_start + offset_col
                widths_dict[col_idx] = w
            set_google_sheet_column_widths(spreadsheet, worksheet.id, widths_dict)
            
        logger.info("Отключенные станции успешно записаны в Google Таблицу.")
    except Exception as e:
        logger.error(f"Не удалось записать отключенные станции в Google Таблицу: {e}")

def send_telegram_document(token, chat_id, file_path, caption):
    """
    Отправляет документ в Telegram.
    
    :param token: str - токен бота
    :param chat_id: str/int - ID чата
    :param file_path: str - путь к отправляемому файлу
    :param caption: str - сообщение к файлу
    :return: bool
    """
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    logger.info(f"Отправка отчета {os.path.basename(file_path)} в Telegram-чат {chat_id}...")
    try:
        with open(file_path, 'rb') as doc:
            files = {'document': doc}
            data = {'chat_id': chat_id, 'caption': caption, 'parse_mode': 'Markdown'}
            response = requests.post(url, data=data, files=files, timeout=30)
            
            if response.status_code == 200:
                logger.info("Отчет успешно отправлен в Telegram!")
                return True
            else:
                logger.error(f"Ошибка отправки Telegram: HTTP {response.status_code}. Ответ: {response.text}")
    except Exception as e:
        logger.error(f"Исключение при отправке Telegram: {e}")
    return False

def make_summary_text(records, city):
    """
    Создает краткий текстовый отчет по количеству задач.
    """
    total_tasks = len(records)
    by_category = {
        'Не в сети': 0,
        'Пустые': 0,
        'Пополнить': 0,
        'Разгрузить': 0,
        'Аппараты с ошибками в ячейках': 0
    }
    for r in records:
        cat = r['Category']
        if cat in by_category:
            by_category[cat] += 1
            
    summary = (
        f"📊 *Отчет по задачам курьеру ({city})*\n"
        f"📅 Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
        f"Всего задач в списке: *{total_tasks}*\n"
        f"🔴 Не в сети: {by_category['Не в сети']}\n"
        f"🟠 Пустые (0 банок): {by_category['Пустые']}\n"
        f"🟡 Пополнить: {by_category['Пополнить']}\n"
        f"🔵 Разгрузить: {by_category['Разгрузить']}\n"
        f"🟣 Ошибки в ячейках: {by_category['Аппараты с ошибками в ячейках']}\n\n"
        f"📎 Сгенерированный файл прикреплен ниже."
    )
    return summary

def main():
    parser = argparse.ArgumentParser(description="Автоматизация обработки задач сервисных инженеров")
    parser.add_argument("--file", help="Путь к конкретному CSV-файлу для обработки")
    parser.add_argument("--test", action="store_true", help="Запуск в тестовом режиме (без отправки в Telegram)")
    parser.add_argument("--download", action="store_true", help="Скачать свежие файлы из Яндекс.Флот перед обработкой")
    parser.add_argument("--headful", action="store_true", help="Запустить браузер в видимом режиме для отладки")
    args = parser.parse_args()

    # Загружаем конфигурацию
    config = load_config()

    # Инициализируем клиент Google Sheets для записи
    gspread_client = authenticate_gspread(config)

    # Создаем директорию для выходных файлов, если её нет
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "outputs")
    inputs_dir = os.path.join(base_dir, "inputs")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    empty_stations_history = load_empty_stations(inputs_dir)

    # 1. Загрузка справочников из Google Sheets
    sheet_id = config.get("google_sheet_id")
    p_sheet = config.get("priority_sheet_name", "test")
    r_sheet = config.get("responsibles_sheet_name", "resp.Omsk")

    df_priorities = fetch_google_sheet(sheet_id, p_sheet)
    df_responsibles = fetch_google_sheet(sheet_id, r_sheet)

    if df_priorities is None or df_responsibles is None:
        logger.error("Критическая ошибка: не удалось получить справочники приоритетов или ответственных!")
        sys.exit(1)

    # Строим словари для быстрого поиска
    # Справочник приоритетов: ID -> приоритет (jewelry)
    # Судя по анализу структуры, первая колонка - vending_id, вторая - jewelry
    priority_col_id = df_priorities.columns[0]
    priority_col_val = df_priorities.columns[1]
    priority_guide = {
        str(int(float(row[priority_col_id]))).strip(): str(row[priority_col_val]).strip()
        for _, row in df_priorities.dropna(subset=[priority_col_id]).iterrows()
    }

    # Справочник ответственных: Номер станции -> Номер контакта
    resp_col_id = df_responsibles.columns[0]
    resp_col_val = df_responsibles.columns[1]
    appointment_person = {
        str(int(float(row[resp_col_id]))).strip(): str(row[resp_col_val]).strip()
        for _, row in df_responsibles.dropna(subset=[resp_col_id]).iterrows()
    }

    # 2. Поиск CSV файлов для обработки
    csv_files = []
    if args.download:
        yandex_parks = config.get("yandex_parks", {})
        if not yandex_parks:
            logger.error("В конфигурации config.json отсутствует раздел 'yandex_parks'!")
            sys.exit(1)
            
        logger.info(f"Начало автоматического скачивания отчетов для парков: {list(yandex_parks.keys())}")
        headless = not args.headful
        
        # Сначала очистим старые файлы в inputs/ для скачиваемых городов
        for park_name in yandex_parks.keys():
            old_file = os.path.join(inputs_dir, f"vendings_{park_name}.csv")
            if os.path.exists(old_file):
                try:
                    os.remove(old_file)
                    logger.info(f"Удален старый файл: {old_file}")
                except Exception as e:
                    logger.warning(f"Не удалось удалить старый файл {old_file}: {e}")
        downloaded_files = []
        for park_name, park_id in yandex_parks.items():
            try:
                dest_path = download_yandex_fleet_csv(config, park_name, park_id, headless=headless)
                downloaded_files.append(os.path.normpath(dest_path))
            except Exception as e:
                logger.error(f"Не удалось скачать данные для парка {park_name}: {e}")
                
        # Будем обрабатывать только те файлы, которые успешно скачались
        csv_files = downloaded_files
    elif args.file:
        if os.path.exists(args.file):
            csv_files.append(args.file)
        else:
            logger.error(f"Указанный файл не найден: {args.file}")
            sys.exit(1)
    else:
        # Сканируем текущую папку и папку inputs на файлы vendings*.csv
        raw_files = glob.glob(os.path.join(base_dir, "vendings*.csv")) + glob.glob(os.path.join(inputs_dir, "vendings*.csv"))
        # Если файлов vendings*.csv нет, проверим любые csv файлы в текущей папке и inputs/
        if not raw_files:
            raw_files = glob.glob(os.path.join(base_dir, "*.csv")) + glob.glob(os.path.join(inputs_dir, "*.csv"))
            raw_files = [f for f in raw_files if "operations" not in os.path.basename(f)]
        
        # Нормализуем пути и удаляем дубликаты
        seen = set()
        for f in raw_files:
            norm_f = os.path.normpath(f)
            if norm_f not in seen:
                seen.add(norm_f)
                csv_files.append(norm_f)

    if not csv_files:
        logger.warning("Не найдено или не скачано ни одного CSV-файла для обработки.")
        sys.exit(0)

    total_files = len(csv_files)
    logger.info(f"Найдено уникальных файлов для обработки: {total_files}")

    for file_idx, file_path in enumerate(csv_files, 1):
        logger.info(f"=== [Файл {file_idx} из {total_files}] Обработка файла: {file_path} ===")
        
        # Определяем кодировку
        encoding = "utf-8"
        try:
            # Пытаемся прочитать первые байты, чтобы угадать UTF-8-SIG (с BOM)
            with open(file_path, 'rb') as f:
                raw = f.read(4)
                if raw.startswith(b'\xef\xbb\xbf'):
                    encoding = 'utf-8-sig'
        except Exception:
            pass

        try:
            df = pd.read_csv(file_path, encoding=encoding)
        except Exception as e:
            logger.error(f"Не удалось прочитать {file_path} с кодировкой {encoding}: {e}")
            logger.info("Пробуем cp1251...")
            try:
                df = pd.read_csv(file_path, encoding="cp1251")
            except Exception as e2:
                logger.error(f"Не удалось прочитать {file_path} с кодировкой cp1251: {e2}")
                continue

        # Определяем основной город из сырого файла до фильтрации для расчета таймзоны
        main_city = "Неизвестно"
        if 'City' in df.columns and not df.empty:
            mode_cities = df['City'].dropna().mode()
            if not mode_cities.empty:
                main_city = str(mode_cities[0]).strip()

        # Загружаем нормы заполненности из сервисной карты для данного города
        map_service_norms = load_map_service_norms(main_city, inputs_dir)

        # Применяем фильтрацию
        records = apply_filtering_logic(df, priority_guide, appointment_person, main_city, map_service_norms, empty_stations_history)
        if not records:
            logger.warning(f"После фильтрации не осталось данных в файле {file_path}")
            continue

        # Обновляем историю пустых станций для обработанного файла
        try:
            # Получаем все уникальные ID автоматов в этом файле
            all_city_vending_ids = set()
            if 'DisplayNumber' in df.columns:
                all_city_vending_ids = set(
                    df['DisplayNumber'].dropna().apply(
                        lambda x: str(int(float(x))).strip() if pd.notna(x) else ""
                    ).unique()
                )
                if "" in all_city_vending_ids:
                    all_city_vending_ids.remove("")

            # Находим те, которые в текущем запуске отмечены как "Пустые"
            current_empty_ids = {r['DisplayNumber'] for r in records if r.get('Category') == 'Пустые'}

            # Рассчитываем локальное время
            CITY_OFFSETS = {
                "омск": 3, "магнитогорск": 2, "сургут": 2, "ижевск": 1, "ульяновск": 1,
                "рязань": 0, "киров": 0, "чебоксары": 0, "орел": 0
            }
            city_key = str(main_city).strip().lower().replace("ё", "е")
            offset = CITY_OFFSETS.get(city_key, 0)
            from datetime import datetime, timezone, timedelta
            local_now = datetime.now(timezone(timedelta(hours=3))) + timedelta(hours=offset)
            local_now_naive = local_now.replace(tzinfo=None)

            for vending_id in all_city_vending_ids:
                if vending_id in current_empty_ids:
                    # Добавляем в историю, если еще нет
                    if vending_id not in empty_stations_history:
                        empty_stations_history[vending_id] = local_now_naive.isoformat()
                else:
                    # Удаляем из истории, если аппарат был пустой, но теперь нет
                    if vending_id in empty_stations_history:
                        empty_stations_history.pop(vending_id, None)
        except Exception as history_update_err:
            logger.error(f"Не удалось обновить историю пустых станций для файла {file_path}: {history_update_err}")

        # Анализируем города в файле и определяем основной город (город-планету)
        df_records = pd.DataFrame(records)
        unique_cities = list(df_records['City'].unique())
        
        # Обновляем основной город по уже отфильтрованным записям для точности
        if 'City' in df_records.columns and not df_records.empty:
            main_city = df_records['City'].mode()[0]
        logger.info(f"В файле обнаружены города: {unique_cities}. Основной город: {main_city}")

        # Генерируем выходное имя файла по основному городу
        safe_city_name = "".join([c for c in main_city if c.isalpha() or c.isspace()]).strip()
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_filename = f"Задачи_{safe_city_name}_{date_str}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        # Сохраняем оформленный Excel со всеми записями (основной город + спутники)
        save_styled_excel(records, output_path)

        # 3. Обновление Google Таблиц
        if gspread_client:
            # Обновление листа основного города в общей таблице задач
            sheet_id_tasks = config.get("google_sheet_id_tasks")
            if sheet_id_tasks:
                update_google_sheet_city(gspread_client, sheet_id_tasks, main_city, records)
            else:
                logger.warning("Параметр google_sheet_id_tasks отсутствует в config.json. Обновление листа города пропущено.")

            # Обновление таблицы отключенных станций ("Не в сети")
            sheet_id_disconnected = config.get("google_sheet_id_disconnected")
            if sheet_id_disconnected:
                update_google_sheet_disconnected(gspread_client, sheet_id_disconnected, main_city, records)
            else:
                logger.warning("Параметр google_sheet_id_disconnected отсутствует в config.json. Обновление отключенных станций пропущено.")
            
            # Небольшая пауза между файлами, чтобы избежать перегрузки лимитов API
            logger.info("Пауза 2 секунды для соблюдения лимитов API...")
            time.sleep(2)

        # 4. Отправка в Telegram чат основного города (ВРЕМЕННО ОТКЛЮЧЕНО по ТЗ)
        # if args.test:
        #     logger.info(f"[ТЕСТ-РЕЖИМ] Пропускаем отправку в Telegram для г. {main_city}")
        # else:
        #     bot_token = config.get("telegram_bot_token")
        #     
        #     # Ищем ID чата регистронезависимо и с обрезкой пробелов
        #     city_chats = config.get("city_chats", {})
        #     chat_id = None
        #     city_clean = str(main_city).strip().lower()
        #     for conf_city, cid in city_chats.items():
        #         if str(conf_city).strip().lower() == city_clean:
        #             chat_id = cid
        #             break
        # 
        #     if not bot_token or bot_token == "YOUR_TELEGRAM_BOT_TOKEN_HERE":
        #         logger.warning("Токен Telegram-бота не заполнен в config.json! Отправка невозможна.")
        #     elif not chat_id or "CHAT_ID_FOR" in str(chat_id):
        #         logger.warning(f"ID чата для города '{main_city}' отсутствует в config.json. Отправка отменена.")
        #     else:
        #         caption = make_summary_text(records, main_city)
        #         send_telegram_document(bot_token, chat_id, output_path, caption)


    # Сохраняем историю пустых станций после завершения всех файлов
    save_empty_stations(inputs_dir, empty_stations_history)

    logger.info("=== Обработка всех файлов завершена! ===")

if __name__ == "__main__":
    main()
