# -*- coding: utf-8 -*-
import os
import sys
import glob
import json
import logging
import argparse
import subprocess
import shutil
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def load_config():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_path = os.path.join(base_dir, 'config.json')
    if not os.path.exists(config_path):
        logger.error(f"Файл конфигурации не найден по пути: {config_path}")
        sys.exit(1)
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def clean_and_parse_numeric(val):
    if pd.isna(val) or val == '':
        return 0.0
    val_str = str(val).strip()
    # Удаляем пробелы и заменяем запятую на точку
    val_str = val_str.replace(' ', '').replace('\xa0', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def process_city_csv(csv_path):
    """
    Парсит CSV выгрузки для города и вычисляет выручку и количество станций.
    
    Args:
        csv_path (str): Абсолютный путь к файлу CSV.
        
    Returns:
        tuple[float, int]: Кортеж из (выручка, количество станций).
    """
    try:
        df = pd.read_csv(csv_path, encoding='utf-8')
    except Exception:
        df = pd.read_csv(csv_path, encoding='cp1251')
        
    # 1. Выручка - сумма всех ячеек столбца "fact"
    df['fact_clean'] = df['fact'].apply(clean_and_parse_numeric)
    total_revenue = df['fact_clean'].sum()
    
    # 2. Кол-во станций - количество строк, где "office_status" = "placed" и в remove_date стоит заглушка (2222-02-01 или 01.02.2222)
    df_active_placed = df[df['office_status'].astype(str).str.strip().str.lower() == 'placed']
    df_active_placed = df_active_placed[df_active_placed['remove_date'].astype(str).str.strip().str.contains('2222-02-01|01.02.2222', regex=True)]
    stations_count = len(df_active_placed)
    
    return total_revenue, stations_count

def find_date_column(sheet, date_row, start_col, end_col, target_date):
    """
    Ищет колонку с целевой датой в указанном диапазоне строк/колонок.
    Если не находит, возвращает первую пустую колонку.
    """
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=date_row, column=col).value
        if isinstance(val, datetime) and val.date() == target_date.date():
            return col
        if isinstance(val, str) and target_date.strftime("%Y-%m-%d") in val:
            return col
            
    # Если точное совпадение не найдено, ищем пустую ячейку в этом диапазоне
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=date_row, column=col).value
        if val is None or str(val).strip() == '':
            sheet.cell(row=date_row, column=col, value=target_date)
            return col
            
    # Если все заполнено, возвращаем следующую за концом диапазона колонку
    new_col = end_col + 1
    sheet.cell(row=date_row, column=new_col, value=target_date)
    return new_col

def main():
    parser = argparse.ArgumentParser(description="Автоматическое заполнение отчета BZ актуальными данными за месяц")
    parser.add_argument("--date", default="2026-06-30", help="Дата отчета (последний день целевого месяца, например 2026-06-30)")
    parser.add_argument("--no-download", action="store_true", help="Пропустить скачивание, использовать имеющиеся файлы в inputs/")
    parser.add_argument("--headful", action="store_true", help="Запустить Playwright в видимом режиме")
    args = parser.parse_args()
    
    config = load_config()
    yandex_parks = config.get("yandex_parks", {})
    
    # Определение целевой даты
    report_date = datetime.strptime(args.date, "%Y-%m-%d")
    target_month_date = report_date.replace(day=1) # Для записи в шапки ячеек
    date_str = report_date.strftime("%Y-%m-%d")
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    inputs_dir = os.path.join(base_dir, "inputs")
    analys_dir = os.path.join(base_dir, "analys")
    debug_dir = os.path.abspath(os.path.join(base_dir, "..", "debug_tools"))
    
    xlsx_files = [f for f in os.listdir(analys_dir) if f.endswith('.xlsx') and 'Годовой_отчет' in f]
    if not xlsx_files:
        logger.error("Excel файл отчета не найден в fleet/analys/!")
        sys.exit(1)
        
    excel_filename = xlsx_files[0]
    excel_path = os.path.join(analys_dir, excel_filename)
    
    logger.info(f"Целевой файл отчета: {excel_path}")
    logger.info(f"Целевой месяц данных: {report_date.strftime('%Y-%m')} (записываемая дата: {target_month_date.strftime('%Y-%m-%d')})")
    
    # 1. Загрузка CSV-отчетов
    csv_paths = {}
    if not args.no_download:
        sys.path.append(base_dir)
        try:
            from yandex_fleet_downloader import download_revenue_report
            logger.info("Запуск скачивания отчетов Yandex Fleet...")
            for park_name, park_id in yandex_parks.items():
                try:
                    csv_path = download_revenue_report(park_name, park_id, report_date, headless=not args.headful)
                    csv_paths[park_name] = csv_path
                except Exception as e:
                    logger.error(f"Не удалось скачать отчет для '{park_name}': {e}. Будет использован локальный файл.")
        except ImportError:
            logger.error("Не удалось импортировать yandex_fleet_downloader. Используются локальные файлы.")
            
    # Поиск файлов в inputs/
    for park_name in yandex_parks.keys():
        if park_name not in csv_paths:
            # Ищем самый свежий файл по паттерну
            pattern = os.path.join(inputs_dir, f"revenue_{park_name}_*.csv")
            matches = glob.glob(pattern)
            if matches:
                # Сортируем по дате изменения
                matches.sort(key=os.path.getmtime)
                csv_paths[park_name] = matches[-1]
                logger.info(f"Для '{park_name}' используем локальный файл: {matches[-1]}")
            else:
                logger.error(f"Файлы отчетов для '{park_name}' отсутствуют в {inputs_dir}!")
                
    # 2. Обработка данных
    city_data = {}
    for city, csv_path in csv_paths.items():
        try:
            rev, st = process_city_csv(csv_path)
            city_data[city] = {
                'revenue': rev,
                'stations': st,
                'rps': rev / st if st > 0 else 0.0
            }
            logger.info(f"Город {city}: Выручка = {rev:,.2f} ₽, Станции = {st}, RPS = {rev/st if st > 0 else 0:,.2f} ₽")
        except Exception as e:
            logger.error(f"Ошибка при обработке CSV для {city}: {e}")
            
    if not city_data:
        logger.error("Нет данных для записи в Excel!")
        sys.exit(1)
        
    # 3. Запись в Excel
    # Используем временный файл во избежание блокировки
    temp_excel_path = os.path.join(debug_dir, 'temp_sync_annual_report.xlsx')
    try:
        subprocess.run(["powershell", "-Command", f'Copy-Item -Path "{excel_path}" -Destination "{temp_excel_path}" -Force'], capture_output=True)
        logger.info("Excel файл скопирован во временную папку")
    except Exception as e:
        logger.warning(f"Ошибка при копировании файла: {e}, пишем напрямую")
        temp_excel_path = excel_path
        
    wb = openpyxl.load_workbook(temp_excel_path, data_only=False)
    
    # 3a. Обновление листа 'Общее'
    if 'Общее' in wb.sheetnames:
        sheet = wb['Общее']
        logger.info("Обновление листа 'Общее'...")
        
        # Города в Excel (названия строк)
        city_rows = {
            'Омск': 3, 'Рязань': 4, 'Ижевск': 5, 'Ульяновск': 6,
            'Магнитогорск': 7, 'Сургут': 8, 'Киров': 9, 'Чебоксары': 10,
            'Орёл': 11
        }
        
        # Нахождение колонок для записи целевого месяца
        # Выручка: строка дат = Row 2
        col_rev = find_date_column(sheet, date_row=2, start_col=2, end_col=49, target_date=target_month_date)
        col_letter = get_column_letter(col_rev)
        
        # Кол-во станций: строка дат = Row 18
        col_st = find_date_column(sheet, date_row=18, start_col=2, end_col=49, target_date=target_month_date)
        
        # Выручка на станцию (RPS): строка дат = Row 57
        col_rps = find_date_column(sheet, date_row=57, start_col=2, end_col=49, target_date=target_month_date)
        
        logger.info(f"Лист 'Общее' -> Колонки записи: Выручка={col_letter} ({col_rev}), Станции={get_column_letter(col_st)}, RPS={get_column_letter(col_rps)}")
        
        # Запись данных по каждому городу
        for city, c_row in city_rows.items():
            if city in city_data:
                data = city_data[city]
                # Выручка
                sheet.cell(row=c_row, column=col_rev, value=data['revenue'])
                # Станции
                # Станции в Excel (Омск=19, Рязань=20, ..., Орёл=27)
                st_row = c_row + 16
                sheet.cell(row=st_row, column=col_st, value=data['stations'])
                # RPS
                # RPS в Excel (Омск=58, Рязань=59, ..., Орёл=66)
                rps_row = c_row + 55
                sheet.cell(row=rps_row, column=col_rps, value=data['rps'])
                
        # Общие формулы (Суммы и Среднее)
        # Сумма Выручки (Row 12)
        sheet.cell(row=12, column=col_rev, value=f"=SUM({col_letter}3:{col_letter}11)")
        # Сумма Станций (Row 28)
        st_letter = get_column_letter(col_st)
        sheet.cell(row=28, column=col_st, value=f"=SUM({st_letter}19:{st_letter}27)")
        # Среднее RPS (Row 67)
        rps_letter = get_column_letter(col_rps)
        sheet.cell(row=67, column=col_rps, value=f"=AVERAGE({rps_letter}58:{rps_letter}66)")
        
    # 3b. Обновление индивидуальных листов городов
    for city in yandex_parks.keys():
        if city in wb.sheetnames and city in city_data:
            sheet = wb[city]
            data = city_data[city]
            logger.info(f"Обновление листа '{city}'...")
            
            # Конфигурация строк для конкретного листа
            if city == 'Орёл':
                row_rev_date, row_rev_val = 1, 2
                row_st_date, row_st_val = 1, 2
                row_rps_date, row_rps_val = 27, 28
                
                # Диапазоны колонок
                col_rev_start, col_rev_end = 1, 10
                col_st_start, col_st_end = 12, 22
                col_rps_start, col_rps_end = 1, 10
            else:
                row_rev_date, row_rev_val = 2, 3
                row_st_date, row_st_val = 2, 3
                row_rps_date, row_rps_val = 47, 48
                
                # Диапазоны колонок
                col_rev_start, col_rev_end = 1, 40
                col_st_start, col_st_end = 31, 70
                col_rps_start, col_rps_end = 1, 40
                
            # Поиск и запись Revenue
            c_rev = find_date_column(sheet, date_row=row_rev_date, start_col=col_rev_start, end_col=col_rev_end, target_date=target_month_date)
            sheet.cell(row=row_rev_val, column=c_rev, value=data['revenue'])
            
            # Поиск и запись Stations
            c_st = find_date_column(sheet, date_row=row_st_date, start_col=col_st_start, end_col=col_st_end, target_date=target_month_date)
            sheet.cell(row=row_st_val, column=c_st, value=data['stations'])
            
            # Поиск и запись RPS
            c_rps = find_date_column(sheet, date_row=row_rps_date, start_col=col_rps_start, end_col=col_rps_end, target_date=target_month_date)
            sheet.cell(row=row_rps_val, column=c_rps, value=data['rps'])
            
    # Сохраняем книгу
    wb.save(temp_excel_path)
    wb.close()
    logger.info("Книга Excel успешно сохранена во временный файл")
    
    # 4. Копирование временного файла обратно
    try:
        subprocess.run(["powershell", "-Command", f'Copy-Item -Path "{temp_excel_path}" -Destination "{excel_path}" -Force'], capture_output=True, check=True)
        logger.info("Отчет успешно обновлен на месте!")
    except Exception as e:
        logger.warning(f"Файл {excel_path} заблокирован (открыт в Excel).")
        base, ext = os.path.splitext(excel_path)
        counter = 1
        saved = False
        while counter <= 100:
            new_path = f"{base}_{counter}{ext}"
            try:
                shutil.copy(temp_excel_path, new_path)
                logger.info(f"Отчет успешно сохранен в альтернативный файл: {new_path}")
                saved = True
                break
            except Exception:
                counter += 1
        if not saved:
            logger.error("Не удалось перезаписать файл отчета.")
            sys.exit(1)

if __name__ == '__main__':
    main()
