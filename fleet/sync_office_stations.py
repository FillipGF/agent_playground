# -*- coding: utf-8 -*-
import os
import sys
import argparse
import logging
import glob
import time
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

# Настройка вывода в консоль
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

def load_config(config_path=None):
    if config_path is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.abspath(os.path.join(base_dir, "..", "config.json"))
        
    if not os.path.exists(config_path):
        logger.error(f"Файл конфигурации не найден: {config_path}")
        sys.exit(1)
    import json
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

from yandex_fleet_downloader import download_revenue_report

def generate_excel_report(data_by_city, output_path):
    """
    Генерирует Excel файл на основе структуры шаблона:
    - Общий лист "Общее" с суммарными данными.
    - Отдельные листы по каждому городу со списками аппаратов за офисом.
    """
    wb = openpyxl.Workbook()
    
    # 1. Лист "Общее"
    ws = wb.active
    ws.title = "Общее"
    
    # Установка ширины столбцов
    ws.column_dimensions['A'].width = 13.71
    ws.column_dimensions['B'].width = 10.43
    ws.column_dimensions['C'].width = 11.43
    ws.column_dimensions['D'].width = 10.43
    
    # Объединение ячеек шапки
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:D2")
    
    # Заполнение шапки
    ws["A1"] = "Город"
    ws["B1"] = "Станции"
    ws["B2"] = "За офисом"
    ws["C2"] = "На локации"
    ws["D1"] = "Всего"
    
    # Центрирование объединенных заголовков
    center_align = Alignment(horizontal="center", vertical="center")
    ws["A1"].alignment = center_align
    ws["B1"].alignment = center_align
    ws["D1"].alignment = center_align
    
    # Заполнение данных по городам
    row_idx = 3
    for city, vals in data_by_city.items():
        ws.cell(row=row_idx, column=1, value=city)
        ws.cell(row=row_idx, column=2, value=vals.get("in_office", 0))
        ws.cell(row=row_idx, column=3, value=vals.get("on_location", 0))
        ws.cell(row=row_idx, column=4, value=f"=SUM(B{row_idx}:C{row_idx})")
        row_idx += 1
        
    # Заполнение строки "Всего"
    total_row = row_idx
    ws.cell(row=total_row, column=1, value="Всего")
    ws.cell(row=total_row, column=2, value=f"=SUM(B3:B{total_row-1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C3:C{total_row-1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{total_row-1})")
    
    # Применение шрифта Calibri 11 ко всем заполненным ячейкам листа "Общее"
    calibri_font = Font(name="Calibri", size=11, bold=False)
    for row in range(1, total_row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = calibri_font
            
    # 2. Отдельные листы по каждому городу
    for city, vals in data_by_city.items():
        # Создаем лист для города
        ws_city = wb.create_sheet(title=city)
        ws_city.views.sheetView[0].showGridLines = True
        
        # Установка ширины столбцов
        ws_city.column_dimensions['A'].width = 13.0
        ws_city.column_dimensions['B'].width = 13.0
        ws_city.column_dimensions['C'].width = 13.0
        ws_city.column_dimensions['D'].width = 18.14
        
        # Заголовки столбцов
        headers = ['DisplayNumber', 'VendingType', 'CellsTotal', 'SerialNumber']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_city.cell(row=1, column=col_idx, value=header)
            cell.font = calibri_font
            
        # Данные аппаратов за офисом
        office_details = vals.get("office_details", [])
        for detail_row_idx, detail in enumerate(office_details, 2):
            ws_city.cell(row=detail_row_idx, column=1, value=detail.get('DisplayNumber'))
            ws_city.cell(row=detail_row_idx, column=2, value=detail.get('VendingType'))
            ws_city.cell(row=detail_row_idx, column=3, value=detail.get('CellsTotal'))
            ws_city.cell(row=detail_row_idx, column=4, value=detail.get('SerialNumber'))
            
            # Применение шрифта ко всем ячейкам строки данных
            for col_idx in range(1, 5):
                ws_city.cell(row=detail_row_idx, column=col_idx).font = calibri_font
                
    # Сохранение файла с обработкой блокировки
    try:
        wb.save(output_path)
        logger.info(f"Excel-отчет успешно сгенерирован и сохранен в: {output_path}")
    except PermissionError:
        logger.warning(f"Файл {output_path} заблокирован (возможно, открыт в Excel).")
        base, ext = os.path.splitext(output_path)
        counter = 1
        saved = False
        while counter <= 100:
            new_path = f"{base}_{counter}{ext}"
            try:
                wb.save(new_path)
                logger.warning(f"Отчет успешно сохранен в альтернативный файл: {new_path}")
                saved = True
                break
            except PermissionError:
                counter += 1
        if not saved:
            logger.error("Не удалось сохранить файл даже под альтернативными именами.")
            raise


def main():
    parser = argparse.ArgumentParser(description="Выгрузка отчетов 'Выручка по аппаратам' и фильтрация станций")
    parser.add_argument("--date", default="yesterday", help="Дата отчета в формате ГГГГ-ММ-ДД или 'yesterday'")
    parser.add_argument("--only-park", help="Скачать/обработать только конкретный город")
    parser.add_argument("--headful", action="store_true", help="Запустить браузер в видимом режиме")
    parser.add_argument("--no-download", action="store_true", help="Пропустить выгрузку, использовать имеющиеся файлы в inputs/")
    args = parser.parse_args()
    
    config = load_config()
    
    # Определение даты отчета (предыдущий день по умолчанию)
    if args.date == "yesterday":
        report_date = datetime.now() - timedelta(days=1)
    else:
        report_date = datetime.strptime(args.date, "%Y-%m-%d")
        
    date_str = report_date.strftime("%Y-%m-%d")
    logger.info(f"Дата выгрузки отчетов: {date_str}")
    
    yandex_parks = config.get("yandex_parks", {})
    if not yandex_parks:
        logger.error("В config.json отсутствует раздел 'yandex_parks'!")
        sys.exit(1)
        
    # Фильтрация по конкретному парку, если передан параметр
    if args.only_park:
        if args.only_park in yandex_parks:
            yandex_parks = {args.only_park: yandex_parks[args.only_park]}
        else:
            logger.error(f"Указанный парк '{args.only_park}' отсутствует в конфигурации!")
            sys.exit(1)
            
    # 1. Скачивание отчетов
    base_dir = os.path.dirname(os.path.abspath(__file__))
    inputs_dir = os.path.join(base_dir, "inputs")
    
    csv_paths = {}
    headless = not args.headful
    
    if args.no_download:
        logger.info(f"Пропуск скачивания. Поиск файлов в папке {inputs_dir}...")
        for park_name in yandex_parks.keys():
            pattern = os.path.join(inputs_dir, f"revenue_{park_name}_{date_str}.csv")
            matches = glob.glob(pattern)
            if matches:
                csv_paths[park_name] = matches[0]
            else:
                # Попробуем найти любые другие файлы этого парка
                fallback_patterns = [
                    os.path.join(inputs_dir, f"revenue_{park_name}_*.csv"),
                    os.path.join(inputs_dir, f"vendings_{park_name}.csv") # самый крайний случай
                ]
                found = False
                for pat in fallback_patterns:
                    matches = glob.glob(pat)
                    if matches:
                        csv_paths[park_name] = matches[0]
                        logger.info(f"Файл за {date_str} не найден для '{park_name}', используем: {matches[0]}")
                        found = True
                        break
                if not found:
                    logger.warning(f"Файлы для '{park_name}' отсутствуют.")
    else:
        for park_name, park_id in yandex_parks.items():
            try:
                csv_path = download_revenue_report(park_name, park_id, report_date, headless=headless)
                csv_paths[park_name] = csv_path
            except Exception as e:
                logger.error(f"Не удалось скачать отчет для '{park_name}': {e}")
                # Вариант А: Ищем старый файл в inputs
                fallback_pattern = os.path.join(inputs_dir, f"revenue_{park_name}_*.csv")
                matches = glob.glob(fallback_pattern)
                if matches:
                    # Берем самый свежий по дате изменения
                    latest_file = max(matches, key=os.path.getmtime)
                    csv_paths[park_name] = latest_file
                    logger.warning(f"Используем ранее скачанный файл: {latest_file}")
                else:
                    logger.error(f"Ранее скачанные файлы для '{park_name}' отсутствуют.")
                    
    # 2. Обработка данных
    data_by_city = {}
    
    # Чтобы сохранить правильный порядок городов из конфига
    for park_name in yandex_parks.keys():
        csv_path = csv_paths.get(park_name)
        if not csv_path or not os.path.exists(csv_path):
            logger.warning(f"Данные по городу {park_name} отсутствуют. Записываем нули.")
            data_by_city[park_name] = {"in_office": 0, "on_location": 0, "office_details": []}
            continue
            
        try:
            df = pd.read_csv(csv_path)
            if 'office_status' not in df.columns:
                logger.error(f"В файле {csv_path} отсутствует столбец 'office_status'!")
                data_by_city[park_name] = {"in_office": 0, "on_location": 0, "office_details": []}
                continue
                
            # Исключаем демонтированные аппараты (оставляем только те, у которых remove_date равен заглушке 2222-02-01 или 01.02.2222)
            if 'remove_date' in df.columns:
                df = df[df['remove_date'].astype(str).str.strip().str.contains('2222-02-01|01.02.2222', regex=True)]
                
            office_col = df['office_status'].fillna('').astype(str).str.lower()
            in_office = int(office_col.str.contains('офис').sum())
            on_location = int(office_col.str.contains('placed').sum())
            
            logger.info(f"Город {park_name}: за офисом = {in_office}, на локации = {on_location}")
            
            # Фильтруем строки для выгрузки за офисом
            df_office = df[office_col.str.contains('офис')].copy()
            office_vending_ids = set(pd.to_numeric(df_office['vending_id'], errors='coerce').dropna().astype(int))
            
            # Пытаемся подгрузить данные из vendings_{park_name}.csv для получения SerialNumber
            vendings_path = os.path.join(inputs_dir, f"vendings_{park_name}.csv")
            office_details = []
            
            if os.path.exists(vendings_path):
                try:
                    df_vend = pd.read_csv(vendings_path)
                    df_vend['DisplayNumber_int'] = pd.to_numeric(df_vend['DisplayNumber'], errors='coerce')
                    
                    # Фильтруем vendings по списку office_vending_ids, сохраняя порядок vendings
                    df_matched = df_vend[df_vend['DisplayNumber_int'].isin(office_vending_ids)].copy()
                    
                    matched_ids = set(df_matched['DisplayNumber_int'].dropna().astype(int))
                    unmatched_ids = office_vending_ids - matched_ids
                    
                    # Добавляем найденные в vendings
                    for _, row in df_matched.iterrows():
                        office_details.append({
                            'DisplayNumber': int(row['DisplayNumber']),
                            'VendingType': row['VendingType'] if pd.notna(row['VendingType']) else "",
                            'CellsTotal': int(row['CellsTotal']) if pd.notna(row['CellsTotal']) else "",
                            'SerialNumber': row['SerialNumber'] if pd.notna(row['SerialNumber']) else ""
                        })
                        
                    # Добавляем ненайденные (fallback на данные из отчета выручки)
                    for um_id in unmatched_ids:
                        rev_row = df_office[pd.to_numeric(df_office['vending_id'], errors='coerce') == um_id].iloc[0]
                        office_details.append({
                            'DisplayNumber': um_id,
                            'VendingType': rev_row.get('model', ""),
                            'CellsTotal': int(rev_row['cells_total']) if pd.notna(rev_row.get('cells_total')) else "",
                            'SerialNumber': ""
                        })
                except Exception as vend_err:
                    logger.error(f"  Ошибка чтения {vendings_path}: {vend_err}. Используем данные без серийных номеров.")
                    # Fallback полностью на выручку
                    for _, row in df_office.iterrows():
                        v_id_val = pd.to_numeric(row['vending_id'], errors='coerce')
                        if pd.notna(v_id_val):
                            office_details.append({
                                'DisplayNumber': int(v_id_val),
                                'VendingType': row.get('model', ""),
                                'CellsTotal': int(row['cells_total']) if pd.notna(row.get('cells_total')) else "",
                                'SerialNumber': ""
                            })
            else:
                logger.warning(f"  Файл {vendings_path} не найден. Данные по аппаратам будут без серийных номеров.")
                # Fallback полностью на выручку
                for _, row in df_office.iterrows():
                    v_id_val = pd.to_numeric(row['vending_id'], errors='coerce')
                    if pd.notna(v_id_val):
                        office_details.append({
                            'DisplayNumber': int(v_id_val),
                            'VendingType': row.get('model', ""),
                            'CellsTotal': int(row['cells_total']) if pd.notna(row.get('cells_total')) else "",
                            'SerialNumber': ""
                        })
                        
            data_by_city[park_name] = {
                "in_office": in_office,
                "on_location": on_location,
                "office_details": office_details
            }
            
        except Exception as e:
            logger.error(f"Ошибка при обработке файла {csv_path}: {e}")
            data_by_city[park_name] = {"in_office": 0, "on_location": 0, "office_details": []}
            
    # 3. Генерация итогового отчета
    outputs_dir = os.path.join(base_dir, "outputs")
    if not os.path.exists(outputs_dir):
        os.makedirs(outputs_dir)
        
    current_date_str = datetime.now().strftime("%d.%m.%Y")
    output_filename = f"Станции за офисом_{current_date_str}.xlsx"
    output_path = os.path.join(outputs_dir, output_filename)
    
    # Также, если мы запускаем только для одного парка, мы можем не перезаписывать общий файл
    # или перезаписать, но для безопасности генерируем всегда
    generate_excel_report(data_by_city, output_path)

if __name__ == "__main__":
    main()
