# -*- coding: utf-8 -*-
import os
import sys
import shutil
import argparse
import logging
import glob
import subprocess
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.formatting.rule import FormulaRule

# Настройка вывода в консоль
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

def copy_cell_style(src_cell, dst_cell):
    """Копирует стили оформления из одной ячейки в другую."""
    if src_cell.has_style:
        dst_cell.font = Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            charset=src_cell.font.charset,
            color=src_cell.font.color,
            underline=src_cell.font.underline,
            strike=src_cell.font.strike
        ) if src_cell.font else None
        dst_cell.fill = PatternFill(
            fill_type=src_cell.fill.fill_type,
            start_color=src_cell.fill.start_color,
            end_color=src_cell.fill.end_color
        ) if src_cell.fill else None
        dst_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        ) if src_cell.border else None
        dst_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            text_rotation=src_cell.alignment.text_rotation,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        ) if src_cell.alignment else None
        dst_cell.number_format = src_cell.number_format

def parse_date_safely(val):
    """Безопасно преобразует значение в дату."""
    if pd.isna(val) or val == '':
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.date()
    val_str = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return val_str

def main():
    parser = argparse.ArgumentParser(description="Синхронизация свода города с данными по выручке за Июнь 2026 года")
    parser.add_argument("--city", default="Омск", help="Название города (например, Омск)")
    parser.add_argument("--date", default="2026-06-30", help="Дата отчета по выручке (ГГГГ-ММ-ДД)")
    args = parser.parse_args()

    city = args.city.strip()
    date_str = args.date.strip()

    # Вычисление относительных путей
    base_dir = os.path.dirname(os.path.abspath(__file__))
    inputs_dir = os.path.join(base_dir, "inputs")
    debug_dir = os.path.abspath(os.path.join(base_dir, "..", "debug_tools"))

    # Поиск файла свода
    svod_pattern = os.path.join(inputs_dir, f"{city}_свод_*.xlsx")
    svod_files = glob.glob(svod_pattern)
    if not svod_files:
        logger.error(f"Файл свода для города '{city}' не найден в {inputs_dir}!")
        sys.exit(1)
    svod_path = svod_files[0]
    logger.info(f"Найден файл свода: {svod_path}")

    # Поиск файла выручки
    revenue_path = os.path.join(inputs_dir, f"revenue_{city}_{date_str}.csv")
    if not os.path.exists(revenue_path):
        # Попробуем найти последний доступный файл
        rev_pattern = os.path.join(inputs_dir, f"revenue_{city}_*.csv")
        rev_files = glob.glob(rev_pattern)
        if not rev_files:
            logger.error(f"Отчет по выручке для города '{city}' не найден в {inputs_dir}!")
            sys.exit(1)
        # Сортируем по имени/дате
        rev_files.sort()
        revenue_path = rev_files[-1]
    logger.info(f"Используется отчет по выручке: {revenue_path}")

    # Путь к временному файлу Excel
    temp_path = os.path.join(debug_dir, f"temp_{city}_свод.xlsx")
    os.makedirs(debug_dir, exist_ok=True)

    # Копирование файла во избежание блокировки
    logger.info(f"Копирование файла свода во временную директорию: {temp_path}")
    try:
        shutil.copy(svod_path, temp_path)
    except PermissionError:
        logger.warning("Permission denied с помощью shutil.copy, пробуем через PowerShell...")
        cmd = f'Copy-Item -Path "{svod_path}" -Destination "{temp_path}" -Force'
        subprocess.run(["powershell", "-Command", cmd], check=True)

    # 1. Загрузка CSV отчета по выручке
    try:
        df_rev = pd.read_csv(revenue_path, encoding='utf-8')
    except Exception:
        df_rev = pd.read_csv(revenue_path, encoding='cp1251')

    # Очистка и приведение типов vending_id
    df_rev = df_rev.dropna(subset=['vending_id'])
    df_rev['vending_id'] = pd.to_numeric(df_rev['vending_id'], errors='coerce').dropna().astype(int)

    # 2. Фильтрация отчета по выручке
    # Оставляем только станции со статусом "placed" (убираем офисы, утерянные и т.д.)
    placed_mask = df_rev['office_status'].fillna('').astype(str).str.strip().str.lower() == 'placed'
    df_rev = df_rev[placed_mask]
    
    # Убираем демонтированные (оставляем только 2222-02-01 или 01.02.2222)
    active_mask = df_rev['remove_date'].astype(str).str.strip().str.contains('2222-02-01|01.02.2222', regex=True)
    df_rev_active = df_rev[active_mask]
    
    logger.info(f"После фильтрации: {len(df_rev_active)} активных аппаратов")

    # Сортируем по place_date по возрастанию, чтобы при наличии дубликатов vending_id
    # в словаре осталась самая свежая запись по дате установки.
    try:
        df_rev_active = df_rev_active.sort_values(by='place_date', ascending=True)
    except Exception:
        pass

    def clean_float(val):
        if pd.isna(val) or val == '':
            return 0.0
        val_str = str(val).strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
        try:
            return float(val_str)
        except ValueError:
            return 0.0

    def clean_int(val):
        if pd.isna(val) or val == '':
            return 0
        val_str = str(val).strip().replace(' ', '').replace('\xa0', '')
        try:
            # handle float-like strings like '3.0' or '3,0'
            val_str = val_str.replace(',', '.')
            if '.' in val_str:
                return int(float(val_str))
            return int(val_str)
        except ValueError:
            return 0

    # Создание словаря активных аппаратов
    active_dict = {}
    for _, row in df_rev_active.iterrows():
        vid = int(row['vending_id'])
        active_dict[vid] = {
            'address': row.get('address', ''),
            'place_name': row.get('place_name', ''),
            'place_date': row.get('place_date', ''),
            'title': row.get('title', ''),
            'fact': clean_float(row.get('fact', 0.0)),
            'orders': clean_int(row.get('orders', 0))
        }

    # 3. Загрузка Excel файла свода
    wb = openpyxl.load_workbook(temp_path, data_only=False)
    
    # Проверка вкладки "Общий"
    sheet_name_svod = "Общий"
    if sheet_name_svod not in wb.sheetnames:
        # Попробуем в нижнем регистре
        found = False
        for name in wb.sheetnames:
            if name.lower() == "общий" or name.lower() == "общее":
                sheet_name_svod = name
                found = True
                break
        if not found:
            logger.error("Лист 'Общий'/'Общее' не найден в Excel!")
            sys.exit(1)

    ws_obsh = wb[sheet_name_svod]
    logger.info(f"Работаем с основным листом: '{ws_obsh.title}'")

    # Поиск колонок "Май, выручка" и "Май, аренды" (последнего года)
    rev_cols = []
    rent_cols = []
    for col in range(1, ws_obsh.max_column + 1):
        val = ws_obsh.cell(row=1, column=col).value
        if val:
            val_str = str(val).strip().replace("\n", " ").replace("  ", " ")
            if "Май" in val_str and "выручка" in val_str:
                rev_cols.append(col)
            elif "Май" in val_str and "аренды" in val_str:
                rent_cols.append(col)

    if not rev_cols or not rent_cols:
        logger.error("Столбцы за 'Май' не найдены в заголовке листа!")
        sys.exit(1)

    target_may_rev = rev_cols[-1]
    target_may_rent = rent_cols[-1]
    logger.info(f"Столбцы Мая (последний блок): выручка = {target_may_rev}, аренды = {target_may_rent}")

    # Вставляем два столбца для Июня сразу после Май, аренды
    june_rev_col = target_may_rent + 1
    june_rent_col = target_may_rent + 2
    logger.info(f"Вставляем новые столбцы Июня на позиции {june_rev_col} и {june_rent_col}")
    ws_obsh.insert_cols(june_rev_col, 2)

    # Записываем заголовки
    ws_obsh.cell(row=1, column=june_rev_col, value="Июнь,\nвыручка")
    ws_obsh.cell(row=1, column=june_rent_col, value="Июнь,\nаренды")

    # Копируем оформление заголовков
    copy_cell_style(ws_obsh.cell(row=1, column=target_may_rev), ws_obsh.cell(row=1, column=june_rev_col))
    copy_cell_style(ws_obsh.cell(row=1, column=target_may_rent), ws_obsh.cell(row=1, column=june_rent_col))

    # Списки для отслеживания обработанных строк
    rows_to_delete = []
    processed_vids = set()

    # Проходим по существующим строкам
    for r in range(2, ws_obsh.max_row + 1):
        v_val = ws_obsh.cell(row=r, column=1).value
        if v_val is None:
            continue
        try:
            vid = int(v_val)
        except ValueError:
            continue

        if vid not in active_dict:
            rows_to_delete.append(r)
        else:
            info = active_dict[vid]
            # Обновление метаданных аппарата
            ws_obsh.cell(row=r, column=2, value=info['address'])
            ws_obsh.cell(row=r, column=3, value=info['place_name'])
            
            # Обработка даты установки
            p_date = parse_date_safely(info['place_date'])
            date_cell = ws_obsh.cell(row=r, column=4, value=p_date)
            if isinstance(p_date, datetime) or hasattr(p_date, "strftime"):
                date_cell.number_format = 'yyyy-mm-dd'
                
            ws_obsh.cell(row=r, column=5, value=info['title'])

            # Запись Июньских данных
            ws_obsh.cell(row=r, column=june_rev_col, value=info['fact'])
            ws_obsh.cell(row=r, column=june_rent_col, value=info['orders'])

            # Копируем оформление ячеек для июня из мая
            copy_cell_style(ws_obsh.cell(row=r, column=target_may_rev), ws_obsh.cell(row=r, column=june_rev_col))
            copy_cell_style(ws_obsh.cell(row=r, column=target_may_rent), ws_obsh.cell(row=r, column=june_rent_col))

            processed_vids.add(vid)

    # Удаление неактивных станций (снизу вверх)
    logger.info(f"Удаление {len(rows_to_delete)} неактивных станций из листа '{ws_obsh.title}'")
    for r in sorted(rows_to_delete, reverse=True):
        ws_obsh.delete_rows(r, 1)

    # Добавление новых станций
    new_vids = set(active_dict.keys()) - processed_vids
    logger.info(f"Добавление {len(new_vids)} новых станций на лист '{ws_obsh.title}'")
    
    for vid in sorted(new_vids):
        info = active_dict[vid]
        new_row = ws_obsh.max_row + 1
        
        # Метаданные
        ws_obsh.cell(row=new_row, column=1, value=vid)
        ws_obsh.cell(row=new_row, column=2, value=info['address'])
        ws_obsh.cell(row=new_row, column=3, value=info['place_name'])
        
        p_date = parse_date_safely(info['place_date'])
        date_cell = ws_obsh.cell(row=new_row, column=4, value=p_date)
        if isinstance(p_date, datetime) or hasattr(p_date, "strftime"):
            date_cell.number_format = 'yyyy-mm-dd'
            
        ws_obsh.cell(row=new_row, column=5, value=info['title'])

        # Копируем стиль с предыдущей строки для всех колонок
        prev_row = new_row - 1
        for col in range(1, june_rent_col + 1):
            copy_cell_style(ws_obsh.cell(row=prev_row, column=col), ws_obsh.cell(row=new_row, column=col))

        # Очищаем исторические значения в скопированных ячейках (колонки 6..target_may_rent)
        for col in range(6, target_may_rent + 1):
            ws_obsh.cell(row=new_row, column=col, value=None)

        # Записываем Июньские данные
        ws_obsh.cell(row=new_row, column=june_rev_col, value=info['fact'])
        ws_obsh.cell(row=new_row, column=june_rent_col, value=info['orders'])

    total_rows = ws_obsh.max_row
    logger.info(f"Итого строк на листе '{ws_obsh.title}': {total_rows}")

    # 4. Обновление Листа 2: "Аналитика новых аппаратов" (с поддержкой опечаток)
    sheet_name_new = None
    for name in wb.sheetnames:
        name_clean = name.lower().replace(" ", "").replace("a", "а").replace("o", "о")
        if "аналитика" in name_clean and ("аппарат" in name_clean or "аппарт" in name_clean):
            sheet_name_new = name
            break
            
    if sheet_name_new:
        ws_new = wb[sheet_name_new]
        logger.info(f"Обновление листа '{sheet_name_new}'...")
        
        # Определяем N в Python
        # Сегодня: 2026-07-07. Граница 5 месяцев: 2026-02-07
        cutoff_date = datetime(2026, 2, 7).date()
        N = 0
        for vid, info in active_dict.items():
            p_date = parse_date_safely(info['place_date'])
            if p_date and isinstance(p_date, (datetime, timedelta, datetime.date.__class__, datetime.strptime.__class__, datetime.now().__class__)) or hasattr(p_date, "year"):
                # handles date objects
                if cutoff_date <= p_date <= datetime(2026, 7, 7).date():
                    N += 1
            elif isinstance(p_date, str) and p_date.strip():
                try:
                    dt = datetime.strptime(p_date.strip()[:10], "%Y-%m-%d").date()
                    if cutoff_date <= dt <= datetime(2026, 7, 7).date():
                        N += 1
                except Exception:
                    pass

        logger.info(f"Количество новых аппаратов на текущую дату (>= 2026-02-07): {N}")

        # Очищаем старые данные
        for r in range(2, ws_new.max_row + 1):
            for c in range(1, ws_new.max_column + 1):
                ws_new.cell(row=r, column=c, value=None)

        # Добавляем новые столбцы шапки Июня, если их там не было
        # Поскольку этот лист повторяет Общий, найдем позицию
        new_rev_cols = []
        new_rent_cols = []
        for col in range(1, ws_new.max_column + 1):
            val = ws_new.cell(row=1, column=col).value
            if val:
                val_str = str(val).strip().replace("\n", " ").replace("  ", " ")
                if "Май" in val_str and "выручка" in val_str:
                    new_rev_cols.append(col)
                elif "Май" in val_str and "аренды" in val_str:
                    new_rent_cols.append(col)

        if new_rev_cols and new_rent_cols:
            t_may_rent = new_rent_cols[-1]
            # Вставляем новые столбцы в шапку листа Аналитики
            ws_new.insert_cols(t_may_rent + 1, 2)
            ws_new.cell(row=1, column=t_may_rent + 1, value="Июнь,\nвыручка")
            ws_new.cell(row=1, column=t_may_rent + 2, value="Июнь,\nаренды")
            
            # Копируем стили
            copy_cell_style(ws_new.cell(row=1, column=new_rev_cols[-1]), ws_new.cell(row=1, column=t_may_rent + 1))
            copy_cell_style(ws_new.cell(row=1, column=t_may_rent), ws_new.cell(row=1, column=t_may_rent + 2))

        # Записываем формулу массива в A2
        # Последняя колонка теперь AY (51-я)
        june_col_letter = openpyxl.utils.get_column_letter(june_rent_col) # AY
        formula_text = f'=FILTER(Общий!A:{june_col_letter}, (Общий!D:D >= DATE(YEAR(TODAY()), MONTH(TODAY())-5, DAY(TODAY()))) * (Общий!D:D <= TODAY()), "Нет данных")'
        # Заворачиваем в префикс для Excel 365
        formula_text_xl = f'=_xlfn._xlws.FILTER(Общий!A:{june_col_letter}, (Общий!D:D >= DATE(YEAR(TODAY()), MONTH(TODAY())-5, DAY(TODAY()))) * (Общий!D:D <= TODAY()), "Нет данных")'
        
        ref_range = f"A2:{june_col_letter}{N+1}" if N > 0 else f"A2:{june_col_letter}2"
        logger.info(f"Запись ArrayFormula на лист Аналитики: {formula_text_xl} с Ref={ref_range}")
        ws_new['A2'] = ArrayFormula(ref_range, formula_text_xl)
        
        # Проставляем '=' в остальные ячейки
        if N > 0:
            for r in range(2, N + 2):
                for c in range(1, june_rent_col + 1):
                    if r == 2 and c == 1:
                        continue
                    ws_new.cell(row=r, column=c, value="=")
                    
        # Копируем оформление ячеек для строк данных из Общего листа
        # Чтобы условное форматирование/границы смотрелись красиво
        for r in range(2, N + 2):
            # берем стиль из аналогичной строки Общего листа (или из строки 2, если Общий длиннее)
            source_row = min(r, ws_obsh.max_row)
            for c in range(1, june_rent_col + 1):
                copy_cell_style(ws_obsh.cell(row=source_row, column=c), ws_new.cell(row=r, column=c))

        # Настройка условного форматирования для колонок Выручки (Май и Июнь)
        logger.info("Применение условного форматирования к колонкам выручки...")
        ws_new.conditional_formatting._cf_rules.clear() # Очищаем старые правила

        green_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        may_rev_letter = openpyxl.utils.get_column_letter(target_may_rev) # 'AV'
        june_rev_letter = openpyxl.utils.get_column_letter(june_rev_col) # 'AX'
        row_limit = N + 1 if N > 0 else 2

        # Правила для Мая
        rule_green_may = FormulaRule(formula=[f'{may_rev_letter}2>=3000'], fill=green_fill)
        rule_yellow_may = FormulaRule(formula=[f'AND({may_rev_letter}2>=1000, {may_rev_letter}2<=2999)'], fill=yellow_fill)
        rule_red_may = FormulaRule(formula=[f'{may_rev_letter}2<=999'], fill=red_fill)

        may_range = f"{may_rev_letter}2:{may_rev_letter}{row_limit}"
        ws_new.conditional_formatting.add(may_range, rule_green_may)
        ws_new.conditional_formatting.add(may_range, rule_yellow_may)
        ws_new.conditional_formatting.add(may_range, rule_red_may)

        # Правила для Июня
        rule_green_june = FormulaRule(formula=[f'{june_rev_letter}2>=3000'], fill=green_fill)
        rule_yellow_june = FormulaRule(formula=[f'AND({june_rev_letter}2>=1000, {june_rev_letter}2<=2999)'], fill=yellow_fill)
        rule_red_june = FormulaRule(formula=[f'{june_rev_letter}2<=999'], fill=red_fill)

        june_range = f"{june_rev_letter}2:{june_rev_letter}{row_limit}"
        ws_new.conditional_formatting.add(june_range, rule_green_june)
        ws_new.conditional_formatting.add(june_range, rule_yellow_june)
        ws_new.conditional_formatting.add(june_range, rule_red_june)

    # 5. Обновление Листа 3: "Просадка выручки" (с поддержкой опечаток)
    sheet_name_pros = None
    for name in wb.sheetnames:
        name_clean = name.lower().replace(" ", "")
        if "просадк" in name_clean:
            sheet_name_pros = name
            break
            
    if sheet_name_pros:
        ws_pros = wb[sheet_name_pros]
        logger.info(f"Обновление листа '{sheet_name_pros}'...")
        
        # Очищаем старые строки данных начиная со 2-й
        if ws_pros.max_row >= 2:
            ws_pros.delete_rows(2, ws_pros.max_row - 1)

        # Копируем шапку Июня
        # Этот лист имеет первый столбец "Проверка", поэтому всё сдвинуто на 1 вправо
        # Найдем Майские столбцы на листе Просадки
        p_rev_cols = []
        p_rent_cols = []
        for col in range(1, ws_pros.max_column + 1):
            val = ws_pros.cell(row=1, column=col).value
            if val:
                val_str = str(val).strip().replace("\n", " ").replace("  ", " ")
                if "Май" in val_str and "выручка" in val_str:
                    p_rev_cols.append(col)
                elif "Май" in val_str and "аренды" in val_str:
                    p_rent_cols.append(col)

        if p_rev_cols and p_rent_cols:
            tp_may_rent = p_rent_cols[-1]
            ws_pros.insert_cols(tp_may_rent + 1, 2)
            ws_pros.cell(row=1, column=tp_may_rent + 1, value="Июнь,\nвыручка")
            ws_pros.cell(row=1, column=tp_may_rent + 2, value="Июнь,\nаренды")
            
            copy_cell_style(ws_pros.cell(row=1, column=p_rev_cols[-1]), ws_pros.cell(row=1, column=tp_may_rent + 1))
            copy_cell_style(ws_pros.cell(row=1, column=tp_may_rent), ws_pros.cell(row=1, column=tp_may_rent + 2))

        # Копируем данные с Общего листа со сдвигом на 1 вправо
        # Колонки Общего листа копируются в столбцы со 2 по 52
        for r_idx in range(2, ws_obsh.max_row + 1):
            # Записываем формулу IF в 1-й столбец
            # AY - Июнь выручка (столбец 51 на листе Просадка), AW - Май выручка (столбец 49 на листе Просадка)
            ws_pros.cell(row=r_idx, column=1, value=f'=IF(AY{r_idx} < AW{r_idx}*0.7, "Да", "Нет")')
            
            # Копируем значения и стили ячеек из Общего листа
            for c_idx in range(1, june_rent_col + 1):
                dst_col = c_idx + 1
                src_cell = ws_obsh.cell(row=r_idx, column=c_idx)
                dst_cell = ws_pros.cell(row=r_idx, column=dst_col, value=src_cell.value)
                copy_cell_style(src_cell, dst_cell)
                
            # Стиль для ячейки формулы в столбце 1 копируем из Май выручки (для единообразия)
            copy_cell_style(ws_obsh.cell(row=r_idx, column=target_may_rev), ws_pros.cell(row=r_idx, column=1))

            # Рассчитываем, просела ли выручка в Python для автофильтра (скрытия строк)
            try:
                june_val = float(ws_obsh.cell(row=r_idx, column=june_rev_col).value or 0)
                may_val = float(ws_obsh.cell(row=r_idx, column=target_may_rev).value or 0)
                is_drop = june_val < may_val * 0.7
            except Exception:
                is_drop = False

            if is_drop:
                ws_pros.row_dimensions[r_idx].hidden = False
            else:
                ws_pros.row_dimensions[r_idx].hidden = True

        # Устанавливаем автофильтр и фильтр по столбцу А (индекс 0) на значение "Да"
        last_col_letter = openpyxl.utils.get_column_letter(june_rent_col + 1)
        ws_pros.auto_filter.ref = f"A1:{last_col_letter}{ws_obsh.max_row}"
        ws_pros.auto_filter.add_filter_column(0, ["Да"])

    # 6. Обновление Листа 4: "Нулевые станции" (с поддержкой опечаток)
    sheet_name_null = None
    for name in wb.sheetnames:
        name_clean = name.lower().replace(" ", "")
        if "нулев" in name_clean:
            sheet_name_null = name
            break
            
    if sheet_name_null:
        ws_null = wb[sheet_name_null]
        logger.info(f"Обновление листа '{sheet_name_null}'...")

        # Вычисляем M в Python на основе данных листа "Общий"
        # Условия: дата установки <= 2026-02-07
        # Выручка в Феврале (AP/42), Марте (AR/44), Апреле (AT/46), Мае (AV/48), Июне (AX/50) < 500
        cutoff_date = datetime(2026, 2, 7).date()
        M = 0
        
        # Мы считываем значения непосредственно из ws_obsh, чтобы посчитать M
        for r in range(2, ws_obsh.max_row + 1):
            p_date_val = ws_obsh.cell(row=r, column=4).value
            p_date = parse_date_safely(p_date_val)
            
            if p_date and (isinstance(p_date, datetime) or hasattr(p_date, "year")):
                # Check installation date
                p_date_only = p_date if not hasattr(p_date, "date") else p_date.date()
                if p_date_only <= cutoff_date:
                    # Check 5 months of revenue
                    # AP (42), AR (44), AT (46), AV (48), AX (50)
                    try:
                        rev_feb = float(ws_obsh.cell(row=r, column=42).value or 0)
                        rev_mar = float(ws_obsh.cell(row=r, column=44).value or 0)
                        rev_apr = float(ws_obsh.cell(row=r, column=46).value or 0)
                        rev_may = float(ws_obsh.cell(row=r, column=48).value or 0)
                        rev_jun = float(ws_obsh.cell(row=r, column=50).value or 0)
                        
                        if rev_feb < 500 and rev_mar < 500 and rev_apr < 500 and rev_may < 500 and rev_jun < 500:
                            M += 1
                    except Exception:
                        pass

        logger.info(f"Количество нулевых станций (установка <= 2026-02-07 и выручка за 5 мес < 500): {M}")

        # Очищаем старые данные
        for r in range(2, ws_null.max_row + 1):
            for c in range(1, ws_null.max_column + 1):
                ws_null.cell(row=r, column=c, value=None)

        # Вставляем June столбцы в шапку листа Нулевых
        n_rev_cols = []
        n_rent_cols = []
        for col in range(1, ws_null.max_column + 1):
            val = ws_null.cell(row=1, column=col).value
            if val:
                val_str = str(val).strip().replace("\n", " ").replace("  ", " ")
                if "Май" in val_str and "выручка" in val_str:
                    n_rev_cols.append(col)
                elif "Май" in val_str and "аренды" in val_str:
                    n_rent_cols.append(col)

        if n_rev_cols and n_rent_cols:
            tn_may_rent = n_rent_cols[-1]
            ws_null.insert_cols(tn_may_rent + 1, 2)
            ws_null.cell(row=1, column=tn_may_rent + 1, value="Июнь,\nвыручка")
            ws_null.cell(row=1, column=tn_may_rent + 2, value="Июнь,\nаренды")
            
            copy_cell_style(ws_null.cell(row=1, column=n_rev_cols[-1]), ws_null.cell(row=1, column=tn_may_rent + 1))
            copy_cell_style(ws_null.cell(row=1, column=tn_may_rent), ws_null.cell(row=1, column=tn_may_rent + 2))

        # Записываем формулу FILTER в A2
        june_col_letter = openpyxl.utils.get_column_letter(june_rent_col) # AY
        formula_text_null_xl = (
            f'=_xlfn._xlws.FILTER(Общий!A2:{june_col_letter}{total_rows},'
            f'(Общий!D2:D{total_rows}<= DATE(YEAR(TODAY()), MONTH(TODAY())-5, DAY(TODAY()))) *'
            f'(Общий!AP2:AP{total_rows}<500)*(Общий!AR2:AR{total_rows}<500)*'
            f'(Общий!AT2:AT{total_rows}<500)*(Общий!AV2:AV{total_rows}<500)*'
            f'(Общий!AX2:AX{total_rows}<500),"Нет данных")'
        )
        
        ref_range_null = f"A2:{june_col_letter}{M+1}" if M > 0 else f"A2:{june_col_letter}2"
        logger.info(f"Запись ArrayFormula на лист Нулевых станций: {formula_text_null_xl} с Ref={ref_range_null}")
        ws_null['A2'] = ArrayFormula(ref_range_null, formula_text_null_xl)

        # Проставляем '=' в остальные ячейки
        if M > 0:
            for r in range(2, M + 2):
                for c in range(1, june_rent_col + 1):
                    if r == 2 and c == 1:
                        continue
                    ws_null.cell(row=r, column=c, value="=")

        # Копируем оформление ячеек для строк данных из Общего листа
        for r in range(2, M + 2):
            source_row = min(r, ws_obsh.max_row)
            for c in range(1, june_rent_col + 1):
                copy_cell_style(ws_obsh.cell(row=source_row, column=c), ws_null.cell(row=r, column=c))

    # 7. Сохранение файла
    wb.save(temp_path)
    wb.close()
    logger.info("Временный файл успешно сохранен!")

    # Генерация имени итогового файла
    months_ru = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель",
        5: "май", 6: "июнь", 7: "июль", 8: "август",
        9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
    }
    now = datetime.now()
    month_name = months_ru[now.month]
    year_val = now.year
    output_filename = f"{city}_{month_name}_{year_val}.xlsx"
    
    outputs_dir = os.path.join(base_dir, "outputs")
    os.makedirs(outputs_dir, exist_ok=True)
    output_path = os.path.join(outputs_dir, output_filename)

    logger.info(f"Сохранение итогового файла: {output_path}")
    try:
        shutil.copy(temp_path, output_path)
        logger.info(f"Итоговый файл успешно сохранен: {output_path}")
    except PermissionError:
        logger.warning(f"Файл {output_path} заблокирован (открыт в Excel).")
        base, ext = os.path.splitext(output_path)
        updated_path = f"{base}_updated{ext}"
        try:
            shutil.copy(temp_path, updated_path)
            logger.info(f"Файл сохранен по альтернативному пути: {updated_path}")
        except Exception as e:
            logger.error(f"Не удалось сохранить даже в альтернативный файл: {e}")
            sys.exit(1)

if __name__ == '__main__':
    main()
