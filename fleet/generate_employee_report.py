# -*- coding: utf-8 -*-
import os
import sys
import argparse
import logging
import glob
import json
import shutil
import subprocess
import calendar
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Настройка вывода в консоль в UTF-8
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Добавляем путь к текущей папке в sys.path для импорта sync_operations
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)
from sync_operations import download_operations_report

def load_config():
    """
    Загружает конфигурационный файл config.json из корня проекта.
    """
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_path = os.path.abspath(os.path.join(base_dir, "config.json"))
    if not os.path.exists(config_path):
        logger.error(f"Файл конфигурации не найден: {config_path}")
        sys.exit(1)
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_copy_file(src, dst):
    """
    Копирует файл с поддержкой копирования заблокированных Excel-файлов на Windows.
    Сначала пытается скопировать стандартным shutil.copy2, а в случае ошибки доступа
    использует PowerShell Copy-Item.
    """
    try:
        shutil.copy2(src, dst)
        return True
    except PermissionError:
        logger.warning(f"Файл {src} заблокирован Excel. Попытка скопировать через PowerShell...")
        cmd = f'powershell -Command "Copy-Item -Path \'{src}\' -Destination \'{dst}\' -Force"'
        res = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        if res.returncode == 0:
            logger.info("Файл успешно скопирован через PowerShell.")
            return True
        else:
            logger.error(f"Не удалось скопировать заблокированный файл. Ошибка: {res.stderr}")
            return False
    except Exception as e:
        logger.error(f"Ошибка при копировании файла {src}: {e}")
        return False

def clean_vending_id(val):
    """
    Очищает ID вендинга, преобразуя в строку с целым числом, если это возможно,
    и отсекая текстовые аномалии типа 'б/н'.
    """
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('б/н', 'б\\н', 'б/н ', 'б\\н ', 'nan', 'none', 'null'):
        return ""
    try:
        return str(int(float(val_str)))
    except ValueError:
        return val_str

def get_latest_excel_form(inputs_dir):
    """
    Находит самый свежий Excel-файл формы в папке inputs/.
    Ищет файлы с расширением .xlsx, исключая временные файлы Excel (начинающиеся с ~$ ).
    """
    pattern = os.path.join(inputs_dir, "*.xlsx")
    files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    # Сортируем по времени изменения
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]

def parse_form_data(form_path, year, month):
    """
    Считывает и парсит данные из формы сервисного инженера за указанный год и месяц.
    """
    logger.info(f"Парсинг формы: {form_path}")
    
    # Делаем временную копию
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    temp_dir = os.path.join(base_dir, "debug_tools")
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, "temp_form_for_report.xlsx")
    
    if not safe_copy_file(form_path, temp_path):
        raise Exception(f"Не удалось открыть файл формы {form_path}")
        
    try:
        df = pd.read_excel(temp_path)
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                logger.warning(f"Не удалось удалить временный файл {temp_path}: {e}")
                
    if df.empty:
        logger.warning("Таблица формы пустая.")
        return pd.DataFrame()
        
    # Проверяем обязательные колонки
    required = ['Время создания', 'Выбери город', 'Номер аппарата', 'Действие']
    for col in required:
        if col not in df.columns:
            raise Exception(f"В форме отсутствует обязательный столбец: {col}")
            
    # Преобразуем даты
    df['Время создания'] = pd.to_datetime(df['Время создания'], errors='coerce')
    
    # Фильтруем строки
    df_filtered = df[
        df['Действие'].notna() &
        df['Время создания'].notna() &
        (df['Время создания'].dt.year == year) &
        (df['Время создания'].dt.month == month) &
        (df['Выбери город'].astype(str).str.strip().str.lower() != 'test')
    ].copy()
    
    tasks = []
    for _, row in df_filtered.iterrows():
        date_val = row['Время создания'].date()
        courier = str(row['Выбери город']).strip()
        vending_id = clean_vending_id(row['Номер аппарата'])
        action = str(row['Действие']).strip()
        
        # Логика выплаты: Пополнение и Выгрузка по 50 р, остальное 100 р
        if action in ("Выгрузка", "Пополнение"):
            payment = 50
        else:
            payment = 100
            
        details = []
        for col_detail in ('Что именно сделал в качестве аккаунтинга?', 'Что сделал в качестве сервисной задачи?', 'Что сделал другое?', 'Комментарий'):
            if col_detail in row and pd.notna(row[col_detail]):
                det_val = str(row[col_detail]).strip()
                if det_val and det_val.lower() not in ('nan', 'none', 'null'):
                    details.append(det_val)
        comment = " | ".join(details)
        
        tasks.append({
            'date': date_val,
            'source': 'Форма',
            'courier': courier,
            'vending_id': vending_id,
            'task_id': '',
            'task_type': action,
            'payment': payment,
            'comment': comment
        })
        
    return pd.DataFrame(tasks)

def parse_yandex_data(inputs_dir, courier_mapping, year, month):
    """
    Считывает и парсит все файлы отчетов по операциям operations_*.csv из папки inputs/ за указанный период.
    """
    logger.info("Парсинг выгрузок Яндекс.Про...")
    csv_files = glob.glob(os.path.join(inputs_dir, "operations_*.csv"))
    
    if not csv_files:
        logger.warning("Файлы operations_*.csv в папке inputs/ не найдены.")
        return pd.DataFrame()
        
    tasks = []
    required_cols = ['ID задачи', 'ФИО исполнителя', 'ID вендинга', 'Тип задачи', 'Итоговый статус', 'Дата итогового статуса', 'Название региона']
    
    for csv_path in csv_files:
        try:
            df = pd.read_csv(csv_path)
            if df.empty:
                continue
                
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                logger.warning(f"Файл {os.path.basename(csv_path)} пропущен: отсутствуют колонки {missing}")
                continue
                
            # Преобразуем даты
            df['Дата итогового статуса'] = pd.to_datetime(df['Дата итогового статуса'], errors='coerce')
            
            # Фильтруем строки за нужный месяц
            df_filtered = df[
                df['Дата итогового статуса'].notna() &
                (df['Дата итогового статуса'].dt.year == year) &
                (df['Дата итогового статуса'].dt.month == month) &
                df['ФИО исполнителя'].notna()
            ].copy()
            
            for _, row in df_filtered.iterrows():
                fio = str(row['ФИО исполнителя']).strip()
                region = str(row['Название региона']).strip()
                task_type = str(row['Тип задачи']).strip()
                status = str(row['Итоговый статус']).strip()
                
                # Загрузка аппарата и Выгрузка аппарата считаются выполненными по умолчанию
                is_loading_unloading = task_type in ("Загрузка аппарата", "Выгрузка аппарата")
                if not is_loading_unloading and status != "Выполнена":
                    continue
                    
                courier_mapped = courier_mapping.get(fio, f"{region} {fio}")
                vending_id = clean_vending_id(row['ID вендинга'])
                task_id = str(row['ID задачи']).strip()
                
                if task_type in ("Загрузка аппарата", "Выгрузка аппарата"):
                    payment = 50
                else:
                    payment = 100
                    
                comment = f"ID задачи: {task_id} | Статус во Fleet: {status}"
                
                tasks.append({
                    'date': row['Дата итогового статуса'].date(),
                    'source': 'Yandex.Pro',
                    'courier': courier_mapped,
                    'vending_id': vending_id,
                    'task_id': task_id,
                    'task_type': task_type,
                    'payment': payment,
                    'comment': comment
                })
        except Exception as e:
            logger.error(f"Ошибка при чтении файла {csv_path}: {e}")
            
    return pd.DataFrame(tasks)

def clean_sheet_name(name):
    r"""
    Создает валидное имя листа Excel на основе ФИО сотрудника.
    Имя должно быть <= 30 символов и не содержать запрещенных знаков: \ / ? * [ ] :
    """
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    cleaned = name
    for char in invalid_chars:
        cleaned = cleaned.replace(char, '')
    return cleaned[:30].strip()

def style_sheet(ws):
    """
    Включает отображение сетки на листе.
    """
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    else:
        ws.sheet_view.showGridLines = True

def write_excel_report(output_path, df_combined, year, month):
    """
    Создает отформатированный Excel-отчет со сводной информацией и детальными горизонтальными матрицами по сотрудникам.
    """
    logger.info(f"Формирование отчета Excel: {output_path}")
    wb = openpyxl.Workbook()
    
    # Количество дней в отчетном месяце
    _, last_day = calendar.monthrange(year, month)
    
    # 1. Лист "Общий расчет"
    ws_summary = wb.active
    ws_summary.title = "Общий расчет"
    style_sheet(ws_summary)
    
    # Заголовок листа
    ws_summary.append([])
    ws_summary.append(["Сводный расчет выплат сотрудникам"])
    ws_summary.cell(2, 1).font = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    
    month_names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    month_name = month_names.get(month, f"Месяц {month}")
    ws_summary.append([f"Период: {month_name} {year} г."])
    ws_summary.cell(3, 1).font = Font(name="Segoe UI", size=11, italic=True)
    ws_summary.append([])
    
    # Шапка таблицы
    headers = ["Сотрудник", "Тип задачи", "Количество задач", "Выплата"]
    ws_summary.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    header_row_idx = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(header_row_idx, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    employees = sorted(df_combined['courier'].unique())
    
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    employee_font = Font(name="Segoe UI", size=11, bold=True)
    detail_font = Font(name="Segoe UI", size=10, italic=False)
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='double', color='000000')
    )
    
    current_row = 6
    
    total_formulas_cnt = []
    total_formulas_pay = []
    
    for emp in employees:
        emp_df = df_combined[df_combined['courier'] == emp]
        
        ws_summary.cell(current_row, 1, emp).font = employee_font
        ws_summary.cell(current_row, 2, "[Итого]").font = employee_font
        
        # Подсчитаем группировку для строки детализации
        type_mapping = {
            "Пополнение": "Загрузка / Пополнение (50 р.)",
            "Загрузка аппарата": "Загрузка / Пополнение (50 р.)",
            "Выгрузка": "Выгрузка (50 р.)",
            "Выгрузка аппарата": "Выгрузка (50 р.)"
        }
        emp_df_mapped = emp_df.copy()
        emp_df_mapped['mapped_type'] = emp_df_mapped['task_type'].apply(lambda x: type_mapping.get(x, f"{x} (100 р.)"))
        
        mapped_breakdown = emp_df_mapped.groupby('mapped_type').agg(
            cnt=('payment', 'count'),
            pay=('payment', 'sum')
        ).reset_index()
        
        detail_count = len(mapped_breakdown)
        detail_start = current_row + 1
        detail_end = current_row + detail_count
        
        # Записываем формулы суммы для строки сотрудника
        ws_summary.cell(current_row, 3, f"=SUM(C{detail_start}:C{detail_end})").font = employee_font
        ws_summary.cell(current_row, 4, f"=SUM(D{detail_start}:D{detail_end})").font = employee_font
        
        total_formulas_cnt.append(f"C{current_row}")
        total_formulas_pay.append(f"D{current_row}")
        
        for col_idx in range(1, 5):
            cell = ws_summary.cell(current_row, col_idx)
            cell.fill = summary_fill
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")
                if col_idx == 4:
                    cell.number_format = '#,##0" р."'
                    
        current_row += 1
        
        # Добавляем детализацию по типам задач
        for _, b_row in mapped_breakdown.iterrows():
            task_type_label = f"  {b_row['mapped_type']}"
            ws_summary.cell(current_row, 2, task_type_label).font = detail_font
            ws_summary.cell(current_row, 3, b_row['cnt']).font = detail_font
            ws_summary.cell(current_row, 4, b_row['pay']).font = detail_font
            
            for col_idx in range(1, 5):
                cell = ws_summary.cell(current_row, col_idx)
                cell.border = thin_border
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal="right")
                    if col_idx == 4:
                        cell.number_format = '#,##0" р."'
            current_row += 1
            
        ws_summary.append([])
        current_row += 1
        
    # Строка общего итога
    ws_summary.cell(current_row, 1, "ИТОГО").font = employee_font
    ws_summary.cell(current_row, 2, "").font = employee_font
    
    grand_cnt_formula = "+" + "+".join(total_formulas_cnt) if total_formulas_cnt else "0"
    grand_pay_formula = "+" + "+".join(total_formulas_pay) if total_formulas_pay else "0"
    
    ws_summary.cell(current_row, 3, f"={grand_cnt_formula}").font = employee_font
    ws_summary.cell(current_row, 4, f"={grand_pay_formula}").font = employee_font
    
    grand_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    for col_idx in range(1, 5):
        cell = ws_summary.cell(current_row, col_idx)
        cell.fill = grand_fill
        cell.border = double_bottom_border
        if col_idx >= 3:
            cell.alignment = Alignment(horizontal="right")
            if col_idx == 4:
                cell.number_format = '#,##0" р."'
                
    # Подгон ширины колонок для сводной
    for col in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 2:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Индивидуальные листы по сотрудникам (ГОРИЗОНТАЛЬНАЯ МАТРИЦА)
    grand_fill_emp = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    for emp in employees:
        emp_df = df_combined[df_combined['courier'] == emp]
        
        sheet_title = clean_sheet_name(emp)
        ws_emp = wb.create_sheet(title=sheet_title)
        style_sheet(ws_emp)
        
        # Шапка листа
        ws_emp.append([])
        ws_emp.append([f"Сводный реестр выполненных задач: {emp}"])
        ws_emp.cell(2, 1).font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
        ws_emp.append([f"Период: {month_name} {year} г."])
        ws_emp.cell(3, 1).font = Font(name="Segoe UI", size=11, italic=True)
        ws_emp.append([])
        
        # Заголовки таблицы матрицы: Тип задачи, дни месяца 01.06..30.06, Всего задач, Тариф, Сумма
        day_headers = [f"{day:02d}.{month:02d}" for day in range(1, last_day + 1)]
        emp_headers = ["Тип задачи"] + day_headers + ["Всего задач", "Тариф", "Сумма"]
        ws_emp.append(emp_headers)
        
        emp_header_row = 5
        for col_idx, header in enumerate(emp_headers, 1):
            cell = ws_emp.cell(emp_header_row, col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Группируем задачи курьера
        type_mapping = {
            "Пополнение": "Загрузка / Пополнение",
            "Загрузка аппарата": "Загрузка / Пополнение",
            "Выгрузка": "Выгрузка",
            "Выгрузка аппарата": "Выгрузка"
        }
        
        emp_df_mapped = emp_df.copy()
        emp_df_mapped['mapped_type'] = emp_df_mapped['task_type'].apply(lambda x: type_mapping.get(x, x))
        emp_df_mapped['day'] = pd.to_datetime(emp_df_mapped['date']).dt.day
        
        def get_tariff(t):
            if t in ("Загрузка / Пополнение", "Выгрузка"):
                return 50
            return 100
            
        unique_types = sorted(emp_df_mapped['mapped_type'].unique())
        
        row_idx = 6
        for t in unique_types:
            tariff = get_tariff(t)
            t_df = emp_df_mapped[emp_df_mapped['mapped_type'] == t]
            
            day_counts = t_df.groupby('day').size().to_dict()
            
            # Заполняем имя типа задачи
            ws_emp.cell(row_idx, 1, t).font = Font(name="Segoe UI", size=10)
            ws_emp.cell(row_idx, 1).border = thin_border
            
            # Количество по каждому дню
            for day in range(1, last_day + 1):
                count = day_counts.get(day, 0)
                cell = ws_emp.cell(row_idx, day + 1)
                if count > 0:
                    cell.value = count
                cell.font = Font(name="Segoe UI", size=10)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                
            # Столбец Всего задач
            last_day_col_let = get_column_letter(last_day + 1)
            cnt_cell = ws_emp.cell(row_idx, last_day + 2, f"=SUM(B{row_idx}:{last_day_col_let}{row_idx})")
            cnt_cell.font = Font(name="Segoe UI", size=10, bold=True)
            cnt_cell.alignment = Alignment(horizontal="center")
            cnt_cell.border = thin_border
            
            # Столбец Тариф
            tariff_cell = ws_emp.cell(row_idx, last_day + 3, tariff)
            tariff_cell.font = Font(name="Segoe UI", size=10)
            tariff_cell.alignment = Alignment(horizontal="right")
            tariff_cell.number_format = '#,##0" р."'
            tariff_cell.border = thin_border
            
            # Столбец Сумма
            sum_col_let = get_column_letter(last_day + 2)
            tariff_col_let = get_column_letter(last_day + 3)
            sum_cell = ws_emp.cell(row_idx, last_day + 4, f"={sum_col_let}{row_idx}*{tariff_col_let}{row_idx}")
            sum_cell.font = Font(name="Segoe UI", size=10, bold=True)
            sum_cell.alignment = Alignment(horizontal="right")
            sum_cell.number_format = '#,##0" р."'
            sum_cell.border = thin_border
            
            row_idx += 1
            
        # Итоговая строка
        total_row_idx = row_idx
        ws_emp.cell(total_row_idx, 1, "Итого за день:").font = employee_font
        ws_emp.cell(total_row_idx, 1).border = thin_border
        
        # Сумма задач по каждому дню
        for day in range(1, last_day + 1):
            col_letter = get_column_letter(day + 1)
            cell = ws_emp.cell(total_row_idx, day + 1, f"=SUM({col_letter}6:{col_letter}{total_row_idx-1})")
            cell.font = employee_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        # Итого количество задач за месяц в столбце Всего задач
        cnt_col_let = get_column_letter(last_day + 2)
        grand_cnt_cell = ws_emp.cell(total_row_idx, last_day + 2, f"=SUM({cnt_col_let}6:{cnt_col_let}{total_row_idx-1})")
        grand_cnt_cell.font = employee_font
        grand_cnt_cell.alignment = Alignment(horizontal="center")
        grand_cnt_cell.border = thin_border
        
        ws_emp.cell(total_row_idx, last_day + 3, "").border = thin_border
        
        # Итого выплата за месяц в столбце Сумма
        pay_col_let = get_column_letter(last_day + 4)
        grand_pay_cell = ws_emp.cell(total_row_idx, last_day + 4, f"=SUM({pay_col_let}6:{pay_col_let}{total_row_idx-1})")
        grand_pay_cell.font = employee_font
        grand_pay_cell.alignment = Alignment(horizontal="right")
        grand_pay_cell.number_format = '#,##0" р."'
        grand_pay_cell.border = thin_border
        
        # Заливка итоговой строки курьера
        for col_idx in range(1, last_day + 5):
            cell = ws_emp.cell(total_row_idx, col_idx)
            cell.fill = grand_fill_emp
            cell.border = double_bottom_border
            
        # Подгон размеров колонок индивидуальных листов
        ws_emp.column_dimensions['A'].width = 28
        for day in range(1, last_day + 1):
            col_letter = get_column_letter(day + 1)
            ws_emp.column_dimensions[col_letter].width = 6
        ws_emp.column_dimensions[get_column_letter(last_day + 2)].width = 13
        ws_emp.column_dimensions[get_column_letter(last_day + 3)].width = 12
        ws_emp.column_dimensions[get_column_letter(last_day + 4)].width = 14
        
    wb.save(output_path)
    logger.info(f"Файл отчета успешно сохранен в: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Объединение задач Яндекс.Про и Формы сервисного инженера в итоговый Excel-отчет")
    parser.add_argument("--month", help="Месяц отчета в формате ГГГГ-ММ (например, 2026-06). Если не указан, берется предыдущий месяц.")
    parser.add_argument("--form-path", help="Путь к Excel-файлу формы. Если не указан, берется самый свежий из папки inputs/.")
    parser.add_argument("--fleet-dir", help="Путь к папке с файлами operations_*.csv отчетов Яндекс.Про. По умолчанию fleet/inputs/.")
    parser.add_argument("--output-path", help="Путь для сохранения итогового файла. По умолчанию сохраняет в fleet/outputs/.")
    parser.add_argument("--download", action="store_true", help="Автоматически скачать актуальные отчеты по операциям из Яндекс.Флит")
    parser.add_argument("--headful", action="store_true", help="Запустить Playwright-браузер для скачивания отчетов в видимом режиме (headful)")
    args = parser.parse_args()

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    inputs_dir = os.path.join(base_dir, "fleet", "inputs")
    outputs_dir = os.path.join(base_dir, "fleet", "outputs")
    os.makedirs(outputs_dir, exist_ok=True)
    
    # 1. Настройка отчетного периода
    if args.month:
        try:
            target_dt = datetime.strptime(args.month, "%Y-%m-%d" if len(args.month) > 7 else "%Y-%m")
            year = target_dt.year
            month = target_dt.month
        except Exception as e:
            logger.error(f"Неверный формат месяца: {args.month}. Используйте формат ГГГГ-ММ. Ошибка: {e}")
            sys.exit(1)
    else:
        now = datetime.now()
        first_day_of_current_month = now.replace(day=1)
        last_day_of_prev_month = first_day_of_current_month - pd.Timedelta(days=1)
        year = last_day_of_prev_month.year
        month = last_day_of_prev_month.month
        
    logger.info(f"Отчетный период: Месяц {month}, Год {year}")
    
    config = load_config()
    
    # 2. Скачивание отчетов из Yandex Fleet при необходимости
    if args.download:
        yandex_parks = config.get("yandex_parks", {})
        if not yandex_parks:
            logger.error("В config.json не заполнен раздел 'yandex_parks'! Скачивание невозможно.")
            sys.exit(1)
            
        logger.info(f"Начало автоматической выгрузки отчетов по операциям из Yandex Fleet за {month}.{year}...")
        
        # Настройка дат выгрузки
        _, last_day = calendar.monthrange(year, month)
        start_dt = datetime(year, month, 1)
        end_dt = datetime(year, month, last_day)
        
        headless = not args.headful
        
        for park_name, park_id in yandex_parks.items():
            try:
                # Импортируем функцию из sync_operations
                logger.info(f"Выгрузка для парка: {park_name}")
                download_operations_report(park_name, park_id, start_dt, end_dt, headless=headless)
            except Exception as e:
                logger.error(f"Не удалось скачать отчет для парка {park_name}: {e}")
                
    # 3. Определение путей к входным данным
    form_path = args.form_path
    if not form_path:
        form_path = get_latest_excel_form(inputs_dir)
        if not form_path:
            logger.error(f"Excel-файл формы не найден в директории {inputs_dir}!")
            sys.exit(1)
    elif not os.path.exists(form_path):
        logger.error(f"Указанный файл формы не найден: {form_path}")
        sys.exit(1)
        
    fleet_dir = args.fleet_dir or inputs_dir
    if not os.path.exists(fleet_dir):
        logger.error(f"Директория отчетов Яндекс.Про не найдена: {fleet_dir}")
        sys.exit(1)
        
    courier_mapping = config.get("courier_mapping", {})
    
    # 4. Парсинг данных
    try:
        df_form = parse_form_data(form_path, year, month)
        logger.info(f"Загружено задач из формы: {len(df_form)}")
    except Exception as e:
        logger.error(f"Ошибка при парсинге формы: {e}")
        sys.exit(1)
        
    try:
        df_yandex = parse_yandex_data(fleet_dir, courier_mapping, year, month)
        logger.info(f"Загружено задач из Яндекс.Про: {len(df_yandex)}")
    except Exception as e:
        logger.error(f"Ошибка при парсинге отчетов Яндекс.Про: {e}")
        sys.exit(1)
        
    # 5. Объединение данных
    dfs_to_concat = []
    if not df_form.empty:
        dfs_to_concat.append(df_form)
    if not df_yandex.empty:
        dfs_to_concat.append(df_yandex)
        
    if not dfs_to_concat:
        logger.error(f"Нет данных по выполненным задачам за указанный период ({month_name} {year}) в обоих источниках!")
        sys.exit(1)
        
    df_combined = pd.concat(dfs_to_concat, ignore_index=True)
    logger.info(f"Всего объединено записей: {len(df_combined)}")
    
    # 6. Запись отчета Excel
    month_str = f"{month:02d}"
    output_path = args.output_path or os.path.join(outputs_dir, f"Расчет_выплат_{year}_{month_str}.xlsx")
    
    try:
        write_excel_report(output_path, df_combined, year, month)
        logger.info("=== Генерация отчета завершена успешно! ===")
    except Exception as e:
        logger.error(f"Не удалось записать итоговый отчет Excel: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
